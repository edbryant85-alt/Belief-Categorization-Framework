from __future__ import annotations

import csv
import json
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from belief_dashboard.export_preview import APPROVED_TO_EVIDENCE
from belief_dashboard.schemas import MI5_COLUMNS, QUEUE_SCHEMAS
from belief_dashboard.utils import timestamp_for_filename, timestamp_iso


APPROVED_COMPARISONS = {
    "evidence_argument": "Evidence / Argument",
    "category": "Category",
    "source_book": "Source / Book",
    "approved_weight_0_5": "Weight 0-5",
    "approved_date": "Date",
    "EC_MI5": "EC MI5",
    "PC_MI5": "PC MI5",
    "PT_MI5": "PT MI5",
    "CT_MI5": "CT MI5",
    "MT_MI5": "MT MI5",
    "IS_MI5": "IS MI5",
    "MS_MI5": "MS MI5",
    "HC_MI5": "HC MI5",
    "N_MI5": "N MI5",
}


def latest_output_workbook(outputs_dir: str | Path) -> Path | None:
    output_path = Path(outputs_dir)
    files = [path for path in output_path.glob("*.xlsx") if path.is_file()]
    if not files:
        return None
    return max(files, key=lambda path: path.stat().st_mtime)


def verify_workbook_export(
    workbook_path: str | Path,
    approved_file: str | Path,
    queue_dir: str | Path,
    config: dict[str, Any],
    *,
    export_report: str | Path | None = None,
    proposal_id: str | None = None,
    source_id: str | None = None,
    mark_exported: bool = False,
    verified_at: datetime | None = None,
) -> dict[str, Any]:
    workbook = Path(workbook_path)
    approved_path = Path(approved_file)
    export_report_path = Path(export_report) if export_report else None
    timestamp = timestamp_iso(verified_at)
    result: dict[str, Any] = {
        "output_workbook_path": str(workbook),
        "approved_updates_file_path": str(approved_path),
        "export_report_path": str(export_report_path) if export_report_path else "",
        "verification_timestamp": timestamp,
        "evidence_log_sheet_name": config["export_verification"]["evidence_log_sheet"],
        "header_row_used": config["export_verification"]["evidence_log_header_row"],
        "approved_rows_considered": 0,
        "matching_exported_rows_found": 0,
        "missing_exported_rows": 0,
        "value_mismatches": 0,
        "formula_concerns": 0,
        "mark_exported_requested": mark_exported,
        "approved_rows_marked_exported": False,
        "rows": [],
        "warnings": [],
        "errors": [],
        "overall_status": "fail",
        "next_step_notes": [],
    }
    if not workbook.exists():
        result["errors"].append(f"Output workbook not found: {workbook}")
        return _finalize(result)
    if not approved_path.exists():
        result["errors"].append(f"Approved updates file not found: {approved_path}")
        return _finalize(result)

    approved_rows = _filter_rows(_read_rows(approved_path), proposal_id=proposal_id, source_id=source_id)
    result["approved_rows_considered"] = len(approved_rows)
    export_report_data = _load_export_report(export_report_path, result) if export_report_path else None

    wb = load_workbook(workbook, read_only=True, data_only=False)
    try:
        sheet_name = result["evidence_log_sheet_name"]
        if sheet_name not in wb.sheetnames:
            result["errors"].append(f"Evidence Log sheet not found: {sheet_name}")
            return _finalize(result)
        sheet = wb[sheet_name]
        headers = _headers_by_name(sheet, int(result["header_row_used"]))
        if "Notes" not in headers:
            result["errors"].append("Evidence Log Notes column not found.")
            return _finalize(result)
        workbook_rows = _rows_by_trace(sheet, headers, int(result["header_row_used"]) + 1)
        formula_columns = _formula_columns(sheet, headers, config)
        for approved in approved_rows:
            row_result = _verify_one_row(approved, workbook_rows, headers, formula_columns, export_report_data)
            result["rows"].append(row_result)
            result["warnings"].extend(row_result["warnings"])
            result["errors"].extend(row_result["errors"])
    finally:
        wb.close()

    result["matching_exported_rows_found"] = sum(1 for row in result["rows"] if row["found"])
    result["missing_exported_rows"] = sum(1 for row in result["rows"] if not row["found"])
    result["value_mismatches"] = sum(len(row["mismatches"]) for row in result["rows"])
    result["formula_concerns"] = sum(len(row["formula_concerns"]) for row in result["rows"])

    if mark_exported and not result["errors"]:
        _mark_approved_rows_exported(
            approved_path,
            {row["proposal_id"] for row in result["rows"]},
            timestamp=timestamp,
            workbook_path=str(workbook),
            report_path="",
            config=config,
        )
        _append_change_log(
            Path(queue_dir) / config["queues"]["files"]["change_log"],
            input_file=str(approved_path),
            output_file=str(workbook),
            status="success",
            details=f"Marked {len(result['rows'])} approved rows exported after verification.",
            changed_at=verified_at,
        )
        result["approved_rows_marked_exported"] = True
    return _finalize(result)


def write_export_verification_reports(
    result: dict[str, Any],
    reports_dir: str | Path,
    *,
    written_at: datetime | None = None,
) -> tuple[Path, Path]:
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)
    stamp = timestamp_for_filename(written_at)
    markdown_path = reports_path / f"export_verification_{stamp}.md"
    json_path = reports_path / f"export_verification_{stamp}.json"
    if result["approved_rows_marked_exported"]:
        _update_export_report_path(
            Path(result["approved_updates_file_path"]),
            {row["proposal_id"] for row in result["rows"]},
            str(json_path),
        )
    markdown_path.write_text(render_export_verification_markdown(result), encoding="utf-8")
    json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    return markdown_path, json_path


def render_export_verification_markdown(result: dict[str, Any]) -> str:
    lines = [
        "# Export Verification Report",
        "",
        f"- Output workbook path: `{result['output_workbook_path']}`",
        f"- Approved updates file path: `{result['approved_updates_file_path']}`",
        f"- Export report path: `{result['export_report_path']}`",
        f"- Verification timestamp: `{result['verification_timestamp']}`",
        f"- Evidence Log sheet name: `{result['evidence_log_sheet_name']}`",
        f"- Header row used: `{result['header_row_used']}`",
        f"- Number of approved rows considered: `{result['approved_rows_considered']}`",
        f"- Number of matching exported rows found: `{result['matching_exported_rows_found']}`",
        f"- Number of missing exported rows: `{result['missing_exported_rows']}`",
        f"- Number of value mismatches: `{result['value_mismatches']}`",
        f"- Number of formula concerns: `{result['formula_concerns']}`",
        f"- Mark exported used: `{result['mark_exported_requested']}`",
        f"- Approved rows marked exported: `{result['approved_rows_marked_exported']}`",
        f"- Overall status: `{result['overall_status']}`",
        "",
        "## Summary",
        "| proposal_id | found | workbook_row | mismatches | formula_concerns |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in result["rows"]:
        lines.append(
            f"| {row['proposal_id']} | {row['found']} | {row['workbook_row']} | "
            f"{len(row['mismatches'])} | {len(row['formula_concerns'])} |"
        )
    lines.extend(["", "## Warnings", *_bullet_list(result["warnings"])])
    lines.extend(["", "## Errors", *_bullet_list(result["errors"])])
    lines.extend(["", "## Next-Step Notes", *_bullet_list(result["next_step_notes"]), ""])
    return "\n".join(lines)


def _verify_one_row(
    approved: dict[str, str],
    workbook_rows: dict[str, dict[str, Any]],
    headers: dict[str, int],
    formula_columns: list[str],
    export_report_data: dict[str, Any] | None,
) -> dict[str, Any]:
    proposal_id = approved.get("proposal_id", "")
    row_result = {
        "proposal_id": proposal_id,
        "claim_id": approved.get("claim_id", ""),
        "source_id": approved.get("source_id", ""),
        "found": False,
        "workbook_row": "",
        "mismatches": [],
        "formula_concerns": [],
        "warnings": [],
        "errors": [],
    }
    workbook_row = workbook_rows.get(proposal_id)
    if not workbook_row:
        row_result["errors"].append(f"{proposal_id}: exported row not found by trace metadata.")
        return row_result
    row_result["found"] = True
    row_result["workbook_row"] = workbook_row["row_number"]
    values = workbook_row["values"]
    for approved_field, workbook_header in APPROVED_COMPARISONS.items():
        expected = approved.get(approved_field, "")
        actual = values.get(workbook_header, "")
        if _norm(expected) != _norm(actual):
            row_result["mismatches"].append(
                f"{workbook_header}: expected '{expected}' but found '{actual}'."
            )
    notes = values.get("Notes", "")
    for token in [
        f"proposal_id={proposal_id}",
        f"claim_id={approved.get('claim_id', '')}",
        f"source_id={approved.get('source_id', '')}",
    ]:
        if token not in notes:
            row_result["errors"].append(f"{proposal_id}: trace metadata missing token {token}.")
    for column in formula_columns:
        if column in headers and _norm(values.get(column, "")) == "":
            row_result["formula_concerns"].append(f"{column} is blank.")
    if export_report_data:
        planned = {
            str(row.get("proposal_id", "")): row
            for row in export_report_data.get("planned_rows", [])
        }.get(proposal_id)
        if planned and str(planned.get("planned_workbook_row", "")) != str(row_result["workbook_row"]):
            row_result["warnings"].append(
                f"{proposal_id}: workbook row differs from export report planned row."
            )
    row_result["errors"].extend(row_result["mismatches"])
    row_result["errors"].extend(row_result["formula_concerns"])
    return row_result


def _headers_by_name(sheet: Any, header_row: int) -> dict[str, int]:
    return {
        str(cell.value).strip(): cell.column
        for cell in sheet[header_row]
        if cell.value is not None and str(cell.value).strip()
    }


def _rows_by_trace(sheet: Any, headers: dict[str, int], start_row: int) -> dict[str, dict[str, Any]]:
    rows: dict[str, dict[str, Any]] = {}
    reverse = {column: header for header, column in headers.items()}
    notes_column = headers["Notes"]
    for row_number in range(start_row, sheet.max_row + 1):
        notes = sheet.cell(row=row_number, column=notes_column).value
        proposal_id = _trace_value(str(notes or ""), "proposal_id")
        if not proposal_id:
            continue
        values = {
            header: sheet.cell(row=row_number, column=column).value
            for column, header in reverse.items()
        }
        rows[proposal_id] = {"row_number": row_number, "values": values}
    return rows


def _formula_columns(sheet: Any, headers: dict[str, int], config: dict[str, Any]) -> list[str]:
    markers = config["workbook_export"]["formula_column_markers"]
    columns = {
        header for header in headers if any(marker.lower() in header.lower() for marker in markers)
    }
    for header, column in headers.items():
        for row_number in range(config["export_verification"]["evidence_log_header_row"] + 1, sheet.max_row + 1):
            value = sheet.cell(row=row_number, column=column).value
            if isinstance(value, str) and value.startswith("="):
                columns.add(header)
                break
    return sorted(columns)


def _mark_approved_rows_exported(
    approved_path: Path,
    proposal_ids: set[str],
    *,
    timestamp: str,
    workbook_path: str,
    report_path: str,
    config: dict[str, Any],
) -> None:
    rows = _read_rows(approved_path)
    for row in rows:
        if row.get("proposal_id") in proposal_ids:
            row["export_status"] = config["export_verification"]["exported_status_label"]
            row["exported_at"] = timestamp
            row["exported_workbook"] = workbook_path
            row["export_verification_report"] = report_path
    _write_rows(approved_path, QUEUE_SCHEMAS["approved_updates"], rows)


def _update_export_report_path(approved_path: Path, proposal_ids: set[str], report_path: str) -> None:
    rows = _read_rows(approved_path)
    for row in rows:
        if row.get("proposal_id") in proposal_ids:
            row["export_verification_report"] = report_path
    _write_rows(approved_path, QUEUE_SCHEMAS["approved_updates"], rows)


def _append_change_log(
    path: Path,
    *,
    input_file: str,
    output_file: str,
    status: str,
    details: str,
    changed_at: datetime | None,
) -> None:
    rows = _read_rows(path)
    rows.append(
        {
            "change_id": f"CHG{len(rows) + 1:04d}",
            "timestamp": (changed_at or datetime.now()).replace(microsecond=0).isoformat(),
            "operation": "verify_workbook_export",
            "input_file": input_file,
            "output_file": output_file,
            "status": status,
            "details": details,
        }
    )
    _write_rows(path, QUEUE_SCHEMAS["change_log"], rows)


def _load_export_report(path: Path | None, result: dict[str, Any]) -> dict[str, Any] | None:
    if path is None:
        return None
    if not path.exists():
        result["errors"].append(f"Export report not found: {path}")
        return None
    return json.loads(path.read_text(encoding="utf-8"))


def _filter_rows(rows: list[dict[str, str]], *, proposal_id: str | None, source_id: str | None) -> list[dict[str, str]]:
    filtered = []
    for row in rows:
        if proposal_id is not None and row.get("proposal_id") != proposal_id:
            continue
        if source_id is not None and row.get("source_id") != source_id:
            continue
        filtered.append(row)
    return filtered


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _write_rows(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    temp_path = path.with_suffix(path.suffix + ".tmp")
    with temp_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})
    os.replace(temp_path, path)


def _trace_value(notes: str, key: str) -> str:
    prefix = f"{key}="
    for part in notes.replace("\n", ";").split(";"):
        value = part.strip()
        if prefix in value:
            return value.split(prefix, 1)[1].strip()
    return ""


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _finalize(result: dict[str, Any]) -> dict[str, Any]:
    if result["errors"]:
        result["overall_status"] = "fail"
        result["next_step_notes"] = ["Fix verification errors before marking rows exported or promoting any workbook."]
    elif result["warnings"]:
        result["overall_status"] = "warning"
        result["next_step_notes"] = ["Review warnings before treating this export as final."]
    else:
        result["overall_status"] = "pass"
        result["next_step_notes"] = ["Export verification passed. The original workbook was not modified."]
    return result


def _bullet_list(items: list[str]) -> list[str]:
    if not items:
        return ["- None"]
    return [f"- {item}" for item in items]
