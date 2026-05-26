from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from belief_dashboard.utils import timestamp_for_filename, timestamp_iso
from belief_dashboard.validation import missing_items, overall_status, present_items


def inspect_workbook(
    workbook_path: str | Path,
    config: dict[str, Any],
    *,
    inspected_at: datetime | None = None,
) -> dict[str, Any]:
    path = Path(workbook_path)
    workbook_config = config["workbook"]
    evidence_config = workbook_config["evidence_log"]
    timestamp = timestamp_iso(inspected_at)

    result: dict[str, Any] = {
        "workbook_path": str(path),
        "inspection_timestamp": timestamp,
        "workbook_file_exists": path.exists(),
        "sheet_names_found": [],
        "expected_sheets": {
            "found": [],
            "missing": list(workbook_config.get("expected_sheets", [])),
        },
        "evidence_log": {
            "sheet_name": evidence_config["sheet_name"],
            "header_row": evidence_config["header_row"],
            "columns_found": [],
            "required_columns": {
                "found": [],
                "missing": list(evidence_config.get("required_columns", [])),
            },
            "hypothesis_mi5_columns": {
                "found": [],
                "missing": [f"{hypothesis} MI5" for hypothesis in workbook_config.get("hypotheses", [])],
            },
            "populated_evidence_rows": 0,
        },
        "overall_status": "fail",
        "next_step_notes": [],
    }

    if not path.exists():
        result["next_step_notes"].append(
            "Workbook file was not found. Place it at the configured path or pass --workbook with an existing .xlsx file."
        )
        return result

    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
        expected_sheets = list(workbook_config.get("expected_sheets", []))
        result["sheet_names_found"] = sheet_names
        result["expected_sheets"] = {
            "found": present_items(expected_sheets, sheet_names),
            "missing": missing_items(expected_sheets, sheet_names),
        }

        evidence_sheet_name = evidence_config["sheet_name"]
        if evidence_sheet_name not in workbook.sheetnames:
            result["next_step_notes"].append(
                f"Evidence Log sheet '{evidence_sheet_name}' was not found."
            )
            result["overall_status"] = overall_status(
                workbook_exists=True,
                missing_expected_sheets=result["expected_sheets"]["missing"],
                missing_required_columns=result["evidence_log"]["required_columns"]["missing"],
                missing_hypothesis_mi5_columns=result["evidence_log"]["hypothesis_mi5_columns"]["missing"],
            )
            return result

        sheet = workbook[evidence_sheet_name]
        header_row_number = int(evidence_config["header_row"])
        headers = _read_header_row(sheet, header_row_number)
        required_columns = list(evidence_config.get("required_columns", []))
        hypothesis_columns = [
            f"{hypothesis} MI5" for hypothesis in workbook_config.get("hypotheses", [])
        ]

        result["evidence_log"]["columns_found"] = headers
        result["evidence_log"]["required_columns"] = {
            "found": present_items(required_columns, headers),
            "missing": missing_items(required_columns, headers),
        }
        result["evidence_log"]["hypothesis_mi5_columns"] = {
            "found": present_items(hypothesis_columns, headers),
            "missing": missing_items(hypothesis_columns, headers),
        }
        result["evidence_log"]["populated_evidence_rows"] = _count_populated_rows(
            sheet,
            start_row=header_row_number + 1,
        )

        result["overall_status"] = overall_status(
            workbook_exists=True,
            missing_expected_sheets=result["expected_sheets"]["missing"],
            missing_required_columns=result["evidence_log"]["required_columns"]["missing"],
            missing_hypothesis_mi5_columns=result["evidence_log"]["hypothesis_mi5_columns"]["missing"],
        )
        result["next_step_notes"] = _next_step_notes(result)
        return result
    finally:
        workbook.close()


def write_reports(
    result: dict[str, Any],
    reports_dir: str | Path,
    *,
    written_at: datetime | None = None,
) -> tuple[Path, Path]:
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)
    stamp = timestamp_for_filename(written_at)
    markdown_path = reports_path / f"workbook_inspection_{stamp}.md"
    json_path = reports_path / f"workbook_inspection_{stamp}.json"

    markdown_path.write_text(render_markdown_report(result), encoding="utf-8")
    json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    return markdown_path, json_path


def render_markdown_report(result: dict[str, Any]) -> str:
    evidence = result["evidence_log"]
    lines = [
        "# Workbook Inspection Report",
        "",
        f"- Workbook path: `{result['workbook_path']}`",
        f"- Inspection timestamp: `{result['inspection_timestamp']}`",
        f"- Workbook file exists: {_yes_no(result['workbook_file_exists'])}",
        f"- Overall status: `{result['overall_status']}`",
        "",
        "## Sheet Names Found",
        *_bullet_list(result["sheet_names_found"]),
        "",
        "## Expected Sheets",
        f"- Found: {_inline_list(result['expected_sheets']['found'])}",
        f"- Missing: {_inline_list(result['expected_sheets']['missing'])}",
        "",
        "## Evidence Log",
        f"- Header row used: `{evidence['header_row']}`",
        f"- Populated evidence rows: `{evidence['populated_evidence_rows']}`",
        "",
        "### Columns Found",
        *_bullet_list(evidence["columns_found"]),
        "",
        "### Required Columns",
        f"- Found: {_inline_list(evidence['required_columns']['found'])}",
        f"- Missing: {_inline_list(evidence['required_columns']['missing'])}",
        "",
        "### Hypothesis MI5 Columns",
        f"- Found: {_inline_list(evidence['hypothesis_mi5_columns']['found'])}",
        f"- Missing: {_inline_list(evidence['hypothesis_mi5_columns']['missing'])}",
        "",
        "## Next-Step Notes",
        *_bullet_list(result["next_step_notes"]),
        "",
    ]
    return "\n".join(lines)


def _read_header_row(sheet: Any, header_row_number: int) -> list[str]:
    values = next(
        sheet.iter_rows(
            min_row=header_row_number,
            max_row=header_row_number,
            values_only=True,
        ),
        (),
    )
    return [str(value).strip() for value in values if value is not None and str(value).strip()]


def _count_populated_rows(sheet: Any, start_row: int) -> int:
    count = 0
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if any(value is not None and str(value).strip() for value in row):
            count += 1
    return count


def _next_step_notes(result: dict[str, Any]) -> list[str]:
    notes = []
    if result["expected_sheets"]["missing"]:
        notes.append("Add or rename missing expected sheets before later phases depend on them.")
    if result["evidence_log"]["required_columns"]["missing"]:
        notes.append("Review the Evidence Log header row and required column configuration.")
    if result["evidence_log"]["hypothesis_mi5_columns"]["missing"]:
        notes.append("Review configured hypotheses or add missing hypothesis MI5 columns.")
    if not notes:
        notes.append("Workbook structure matches Phase 1 expectations. No workbook changes were made.")
    return notes


def _bullet_list(items: list[str]) -> list[str]:
    if not items:
        return ["- None"]
    return [f"- {item}" for item in items]


def _inline_list(items: list[str]) -> str:
    if not items:
        return "None"
    return ", ".join(f"`{item}`" for item in items)


def _yes_no(value: bool) -> str:
    return "yes" if value else "no"
