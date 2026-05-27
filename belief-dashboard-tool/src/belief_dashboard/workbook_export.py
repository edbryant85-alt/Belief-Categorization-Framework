from __future__ import annotations

import csv
import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

from belief_dashboard.export_preview import preview_workbook_export
from belief_dashboard.schemas import QUEUE_SCHEMAS
from belief_dashboard.utils import timestamp_for_filename, timestamp_iso


CHANGE_LOG_HEADERS = QUEUE_SCHEMAS["change_log"]

PLAN_TO_WORKBOOK = {
    "planned_evidence_id": "ID",
    "date": "Date",
    "evidence_argument": "Evidence / Argument",
    "category": "Category",
    "source_book": "Source / Book",
    "weight_0_5": "Weight 0-5",
    "EC_MI5": "EC MI5",
    "PC_MI5": "PC MI5",
    "PT_MI5": "PT MI5",
    "CT_MI5": "CT MI5",
    "MT_MI5": "MT MI5",
    "IS_MI5": "IS MI5",
    "MS_MI5": "MS MI5",
    "HC_MI5": "HC MI5",
    "N_MI5": "N MI5",
    "notes": "Notes",
}


def apply_approved_to_workbook(
    workbook_path: str | Path,
    approved_file: str | Path,
    queue_dir: str | Path,
    config: dict[str, Any],
    *,
    backups_dir: str | Path,
    outputs_dir: str | Path,
    dry_run: bool = False,
    limit: int | None = None,
    proposal_id: str | None = None,
    source_id: str | None = None,
    applied_at: datetime | None = None,
) -> dict[str, Any]:
    timestamp = timestamp_iso(applied_at)
    stamp = timestamp_for_filename(applied_at)
    workbook = Path(workbook_path)
    backup_path = Path(backups_dir) / _backup_name(workbook, stamp)
    output_path = Path(outputs_dir) / _output_name(workbook, stamp)
    preview = preview_workbook_export(
        workbook,
        approved_file,
        queue_dir,
        config,
        limit=limit,
        proposal_id=proposal_id,
        source_id=source_id,
        previewed_at=applied_at,
    )
    result = {
        "workbook_path": str(workbook),
        "backup_workbook_path": "" if dry_run else str(backup_path),
        "output_workbook_path": "" if dry_run else str(output_path),
        "approved_updates_file_path": str(approved_file),
        "export_timestamp": timestamp,
        "dry_run": dry_run,
        "evidence_log_sheet_name": preview["evidence_log_sheet_name"],
        "header_row_used": preview["header_row_used"],
        "existing_populated_evidence_row_count": preview["existing_populated_evidence_row_count"],
        "first_appended_row": preview["first_planned_append_row"],
        "approved_rows_considered": preview["approved_rows_considered"],
        "rows_exported": 0,
        "rows_blocked": preview["rows_blocked_by_validation_errors"],
        "workbook_input_columns": preview["workbook_input_columns"],
        "formula_driven_columns_copied": [],
        "formula_driven_columns_skipped_or_warned": [],
        "id_planning_status": preview["id_planning_status"],
        "planned_rows": preview["planned_rows"],
        "warnings": list(preview["warnings"]),
        "errors": list(preview["errors"]),
        "overall_status": "fail",
        "next_step_notes": [],
        "backup_created": False,
        "output_created": False,
        "change_log_updated": False,
    }
    if preview["overall_status"] == "fail":
        return _finalize(result, exported=False)
    if dry_run:
        result["rows_exported"] = preview["rows_ready_for_export"]
        return _finalize(result, exported=False)
    if output_path.exists() or backup_path.exists():
        result["errors"].append("Backup or output workbook already exists; refusing to overwrite.")
        return _finalize(result, exported=False)
    if preview["rows_ready_for_export"] == 0:
        result["warnings"].append("No approved rows were ready for export.")

    backup_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(workbook, backup_path)
    result["backup_created"] = True
    shutil.copy2(workbook, output_path)

    try:
        copied, skipped = _write_output_workbook(output_path, preview, config)
    except Exception as exc:
        if output_path.exists():
            output_path.unlink()
        result["errors"].append(f"Workbook export failed before saving output workbook: {exc}")
        return _finalize(result, exported=False)

    result["formula_driven_columns_copied"] = copied
    result["formula_driven_columns_skipped_or_warned"] = skipped
    result["rows_exported"] = preview["rows_ready_for_export"]
    result["output_created"] = True
    _append_change_log(
        Path(queue_dir) / config["queues"]["files"]["change_log"],
        input_file=str(approved_file),
        output_file=str(output_path),
        status="success",
        details=f"Applied {result['rows_exported']} approved rows to timestamped workbook copy.",
        changed_at=applied_at,
    )
    result["change_log_updated"] = True
    return _finalize(result, exported=True)


def write_workbook_export_report(
    result: dict[str, Any],
    reports_dir: str | Path,
    *,
    written_at: datetime | None = None,
) -> tuple[Path, Path]:
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)
    stamp = timestamp_for_filename(written_at)
    markdown_path = reports_path / f"workbook_export_{stamp}.md"
    json_path = reports_path / f"workbook_export_{stamp}.json"
    markdown_path.write_text(render_workbook_export_markdown(result), encoding="utf-8")
    json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    return markdown_path, json_path


def render_workbook_export_markdown(result: dict[str, Any]) -> str:
    lines = [
        "# Workbook Export Report",
        "",
        f"- Workbook path: `{result['workbook_path']}`",
        f"- Backup workbook path: `{result['backup_workbook_path']}`",
        f"- Output workbook path: `{result['output_workbook_path']}`",
        f"- Approved updates file path: `{result['approved_updates_file_path']}`",
        f"- Export timestamp: `{result['export_timestamp']}`",
        f"- Dry-run status: `{result['dry_run']}`",
        f"- Evidence Log sheet name: `{result['evidence_log_sheet_name']}`",
        f"- Header row used: `{result['header_row_used']}`",
        f"- Existing populated evidence row count: `{result['existing_populated_evidence_row_count']}`",
        f"- First appended row: `{result['first_appended_row']}`",
        f"- Approved rows considered: `{result['approved_rows_considered']}`",
        f"- Rows exported: `{result['rows_exported']}`",
        f"- Rows blocked: `{result['rows_blocked']}`",
        f"- ID planning status: `{result['id_planning_status']}`",
        f"- Overall status: `{result['overall_status']}`",
        "",
        "## Workbook Input Columns",
        f"- Found: {_inline(result['workbook_input_columns']['found'])}",
        f"- Missing: {_inline(result['workbook_input_columns']['missing'])}",
        "",
        "## Formula Columns Copied",
        *_bullet_list(result["formula_driven_columns_copied"]),
        "",
        "## Formula Columns Skipped Or Warned",
        *_bullet_list(result["formula_driven_columns_skipped_or_warned"]),
        "",
        "## Exported Or Planned Rows",
        "| workbook_row | evidence_id | proposal_id | status | warnings |",
        "| --- | --- | --- | --- | --- |",
    ]
    for row in result["planned_rows"]:
        lines.append(
            f"| {row['planned_workbook_row']} | {row['planned_evidence_id']} | "
            f"{row['proposal_id']} | {row['validation_status']} | {row['warnings']} |"
        )
    lines.extend(["", "## Warnings", *_bullet_list(result["warnings"])])
    lines.extend(["", "## Errors", *_bullet_list(result["errors"])])
    lines.extend(["", "## Next-Step Notes", *_bullet_list(result["next_step_notes"]), ""])
    return "\n".join(lines)


def _write_output_workbook(output_path: Path, preview: dict[str, Any], config: dict[str, Any]) -> tuple[list[str], list[str]]:
    workbook = load_workbook(output_path)
    sheet = workbook[preview["evidence_log_sheet_name"]]
    headers = _headers_by_name(sheet, int(preview["header_row_used"]))
    formula_headers = [
        header for header in preview["formula_driven_columns_detected"] if header in headers
    ]
    copied: set[str] = set()
    skipped: set[str] = set()
    for plan_row in preview["planned_rows"]:
        if plan_row["validation_status"] != "ready":
            continue
        target_row = int(plan_row["planned_workbook_row"])
        source_formula_row = target_row - 1
        for plan_field, workbook_header in PLAN_TO_WORKBOOK.items():
            if workbook_header in headers:
                sheet.cell(row=target_row, column=headers[workbook_header], value=plan_row.get(plan_field, ""))
        if config["workbook_export"].get("copy_formulas_down", True):
            for header in formula_headers:
                column = headers[header]
                formula_row = _nearest_formula_row(sheet, column, source_formula_row, int(preview["header_row_used"]) + 1)
                if formula_row is None:
                    skipped.add(header)
                    continue
                source_cell = sheet.cell(row=formula_row, column=column)
                target_cell = sheet.cell(row=target_row, column=column)
                target_cell.value = Translator(source_cell.value, origin=source_cell.coordinate).translate_formula(target_cell.coordinate)
                copied.add(header)
    workbook.save(output_path)
    workbook.close()
    return sorted(copied), sorted(skipped)


def _headers_by_name(sheet: Any, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[header_row]:
        if cell.value is not None and str(cell.value).strip():
            headers[str(cell.value).strip()] = cell.column
    return headers


def _nearest_formula_row(sheet: Any, column: int, start_row: int, min_row: int) -> int | None:
    for row_number in range(start_row, min_row - 1, -1):
        value = sheet.cell(row=row_number, column=column).value
        if isinstance(value, str) and value.startswith("="):
            return row_number
    return None


def _append_change_log(
    path: Path,
    *,
    input_file: str,
    output_file: str,
    status: str,
    details: str,
    changed_at: datetime | None,
) -> None:
    rows = _read_rows(path)
    rows.append(
        {
            "change_id": f"CHG{len(rows) + 1:04d}",
            "timestamp": (changed_at or datetime.now()).replace(microsecond=0).isoformat(),
            "operation": "apply_approved_to_workbook",
            "input_file": input_file,
            "output_file": output_file,
            "status": status,
            "details": details,
        }
    )
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CHANGE_LOG_HEADERS)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in CHANGE_LOG_HEADERS})


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _backup_name(workbook: Path, stamp: str) -> str:
    return f"{workbook.stem}_backup_{stamp}{workbook.suffix}"


def _output_name(workbook: Path, stamp: str) -> str:
    return f"{workbook.stem}_exported_{stamp}{workbook.suffix}"


def _finalize(result: dict[str, Any], *, exported: bool) -> dict[str, Any]:
    if result["errors"]:
        result["overall_status"] = "fail"
        result["next_step_notes"] = ["Fix validation or workbook export errors before trying again."]
    elif result["warnings"]:
        result["overall_status"] = "warning"
        result["next_step_notes"] = ["Export completed or previewed with warnings. Review the report before using the output workbook."]
    else:
        result["overall_status"] = "pass"
        result["next_step_notes"] = ["Workbook export completed successfully." if exported else "Dry run completed successfully. No workbook files were written."]
    return result


def _bullet_list(items: list[str]) -> list[str]:
    if not items:
        return ["- None"]
    return [f"- {item}" for item in items]


def _inline(items: list[str]) -> str:
    if not items:
        return "None"
    return ", ".join(f"`{item}`" for item in items)
