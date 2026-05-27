from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from belief_dashboard.schemas import MI5_COLUMNS
from belief_dashboard.utils import timestamp_for_filename, timestamp_iso


APPROVED_TO_EVIDENCE = {
    "evidence_argument": "Evidence / Argument",
    "category": "Category",
    "source_book": "Source / Book",
    "approved_weight_0_5": "Weight 0-5",
    "EC_MI5": "EC MI5",
    "PC_MI5": "PC MI5",
    "PT_MI5": "PT MI5",
    "CT_MI5": "CT MI5",
    "MT_MI5": "MT MI5",
    "IS_MI5": "IS MI5",
    "MS_MI5": "MS MI5",
    "HC_MI5": "HC MI5",
    "N_MI5": "N MI5",
    "approved_date": "Date",
}

CHANGE_PLAN_HEADERS = [
    "planned_workbook_row",
    "planned_evidence_id",
    "proposal_id",
    "claim_id",
    "source_id",
    "evidence_argument",
    "category",
    "source_book",
    "weight_0_5",
    "EC_MI5",
    "PC_MI5",
    "PT_MI5",
    "CT_MI5",
    "MT_MI5",
    "IS_MI5",
    "MS_MI5",
    "HC_MI5",
    "N_MI5",
    "date",
    "notes",
    "validation_status",
    "warnings",
]


def preview_workbook_export(
    workbook_path: str | Path,
    approved_file: str | Path,
    queue_dir: str | Path,
    config: dict[str, Any],
    *,
    limit: int | None = None,
    proposal_id: str | None = None,
    source_id: str | None = None,
    previewed_at: datetime | None = None,
) -> dict[str, Any]:
    workbook = Path(workbook_path)
    approved = Path(approved_file)
    queue_path = Path(queue_dir)
    export_config = config["workbook_export"]
    timestamp = timestamp_iso(previewed_at)
    result: dict[str, Any] = {
        "workbook_path": str(workbook),
        "approved_updates_file_path": str(approved),
        "inspection_timestamp": timestamp,
        "evidence_log_sheet_name": export_config["evidence_log_sheet"],
        "header_row_used": export_config["evidence_log_header_row"],
        "existing_populated_evidence_row_count": 0,
        "first_planned_append_row": None,
        "approved_rows_considered": 0,
        "rows_ready_for_export": 0,
        "rows_blocked_by_validation_errors": 0,
        "workbook_input_columns": {"found": [], "missing": []},
        "formula_driven_columns_detected": [],
        "id_planning_status": "not_checked",
        "planned_rows": [],
        "warnings": [],
        "errors": [],
        "overall_status": "fail",
        "next_step_notes": [],
    }
    if not workbook.exists():
        result["errors"].append(f"Workbook file not found: {workbook}")
        return _finalize(result)
    if not approved.exists():
        result["errors"].append(f"Approved updates file not found: {approved}")
        return _finalize(result)

    wb = load_workbook(workbook, read_only=True, data_only=False)
    try:
        sheet_name = export_config["evidence_log_sheet"]
        if sheet_name not in wb.sheetnames:
            result["errors"].append(f"Evidence Log sheet not found: {sheet_name}")
            return _finalize(result)
        sheet = wb[sheet_name]
        headers = _read_header_row(sheet, int(export_config["evidence_log_header_row"]))
        required_columns = list(export_config["input_columns"].values())
        result["workbook_input_columns"] = {
            "found": [column for column in required_columns if column in headers],
            "missing": [column for column in required_columns if column not in headers],
        }
        formula_columns = _detect_formula_columns(sheet, headers, config)
        result["formula_driven_columns_detected"] = formula_columns
        header_row = int(export_config["evidence_log_header_row"])
        first_data_row = header_row + 1
        meaningful = _meaningful_evidence_rows(sheet, headers, first_data_row, config)
        result["existing_populated_evidence_row_count"] = len(meaningful)
        last_meaningful_row = max(meaningful) if meaningful else header_row
        result["first_planned_append_row"] = last_meaningful_row + 1
        id_plan = _plan_ids(sheet, headers, first_data_row, last_meaningful_row)
        result["id_planning_status"] = id_plan["status"]
        result["warnings"].extend(id_plan["warnings"])
    finally:
        wb.close()

    if result["workbook_input_columns"]["missing"]:
        result["errors"].append(
            "Workbook is missing required Evidence Log input columns: "
            + ", ".join(result["workbook_input_columns"]["missing"])
        )

    approved_rows = _filter_rows(_read_csv_rows(approved), proposal_id=proposal_id, source_id=source_id, limit=limit)
    result["approved_rows_considered"] = len(approved_rows)
    refs = _load_reference_ids(queue_path, config)
    next_ids = _planned_ids_for_rows(result["id_planning_status"], id_plan.get("next_id"), len(approved_rows))
    for index, row in enumerate(approved_rows):
        planned_row = _build_change_plan_row(
            row,
            planned_workbook_row=(result["first_planned_append_row"] or 0) + index,
            planned_evidence_id=next_ids[index],
            config=config,
            refs=refs,
        )
        result["planned_rows"].append(planned_row)

    result["rows_ready_for_export"] = sum(1 for row in result["planned_rows"] if row["validation_status"] == "ready")
    result["rows_blocked_by_validation_errors"] = len(result["planned_rows"]) - result["rows_ready_for_export"]
    for row in result["planned_rows"]:
        result["errors"].extend(row.pop("_errors"))
    return _finalize(result)


def write_export_preview_artifacts(
    result: dict[str, Any],
    reports_dir: str | Path,
    *,
    written_at: datetime | None = None,
) -> tuple[Path, Path, Path]:
    reports_path = Path(reports_dir)
    reports_path.mkdir(parents=True, exist_ok=True)
    stamp = timestamp_for_filename(written_at)
    markdown_path = reports_path / f"workbook_export_preview_{stamp}.md"
    json_path = reports_path / f"workbook_export_preview_{stamp}.json"
    csv_path = reports_path / f"workbook_export_change_plan_{stamp}.csv"
    markdown_path.write_text(render_export_preview_markdown(result), encoding="utf-8")
    json_path.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8")
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CHANGE_PLAN_HEADERS)
        writer.writeheader()
        for row in result["planned_rows"]:
            writer.writerow({header: row.get(header, "") for header in CHANGE_PLAN_HEADERS})
    return markdown_path, json_path, csv_path


def render_export_preview_markdown(result: dict[str, Any]) -> str:
    lines = [
        "# Workbook Export Preview",
        "",
        f"- Workbook path: `{result['workbook_path']}`",
        f"- Approved updates file path: `{result['approved_updates_file_path']}`",
        f"- Inspection timestamp: `{result['inspection_timestamp']}`",
        f"- Evidence Log sheet name: `{result['evidence_log_sheet_name']}`",
        f"- Header row used: `{result['header_row_used']}`",
        f"- Existing populated evidence row count: `{result['existing_populated_evidence_row_count']}`",
        f"- First planned append row: `{result['first_planned_append_row']}`",
        f"- Number of approved rows considered: `{result['approved_rows_considered']}`",
        f"- Number of rows ready for export: `{result['rows_ready_for_export']}`",
        f"- Number of rows blocked by validation errors: `{result['rows_blocked_by_validation_errors']}`",
        f"- ID planning status: `{result['id_planning_status']}`",
        f"- Overall status: `{result['overall_status']}`",
        "",
        "## Workbook Input Columns",
        f"- Found: {_inline(result['workbook_input_columns']['found'])}",
        f"- Missing: {_inline(result['workbook_input_columns']['missing'])}",
        "",
        "## Formula-Driven Columns Detected",
        *_bullet_list(result["formula_driven_columns_detected"]),
        "",
        "## Planned Rows",
        "| workbook_row | evidence_id | proposal_id | source_id | claim_id | status | warnings |",
        "| --- | --- | --- | --- | --- | --- | --- |",
    ]
    for row in result["planned_rows"]:
        lines.append(
            f"| {row['planned_workbook_row']} | {row['planned_evidence_id']} | {row['proposal_id']} | "
            f"{row['source_id']} | {row['claim_id']} | {row['validation_status']} | {row['warnings']} |"
        )
    lines.extend(["", "## Warnings", *_bullet_list(result["warnings"])])
    lines.extend(["", "## Errors", *_bullet_list(result["errors"])])
    lines.extend(["", "## Next-Step Notes", *_bullet_list(result["next_step_notes"]), ""])
    return "\n".join(lines)


def _build_change_plan_row(
    row: dict[str, str],
    *,
    planned_workbook_row: int,
    planned_evidence_id: str,
    config: dict[str, Any],
    refs: dict[str, set[str]],
) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    for field in ["proposal_id", "claim_id", "source_id", "evidence_argument", "category", "source_book", "approved_weight_0_5", "approved_by", "approved_date"]:
        if not (row.get(field) or "").strip():
            errors.append(f"{row.get('proposal_id') or '(blank proposal_id)'}: {field} is required.")
    _validate_weight(row.get("approved_weight_0_5", ""), row.get("proposal_id", ""), errors)
    _validate_mi5(row, config, errors)
    _validate_date(row.get("approved_date", ""), row.get("proposal_id", ""), errors)
    if row.get("proposal_id", "") not in refs["proposal_ids"]:
        errors.append(f"{row.get('proposal_id')}: proposal_id not found in proposed_updates.csv.")
    if row.get("claim_id", "") not in refs["claim_ids"]:
        errors.append(f"{row.get('proposal_id')}: claim_id not found in extracted_claims.csv.")
    if row.get("source_id", "") not in refs["source_ids"]:
        errors.append(f"{row.get('proposal_id')}: source_id not found in source_dossiers.csv.")
    notes = _combined_notes(row)
    return {
        "planned_workbook_row": planned_workbook_row,
        "planned_evidence_id": planned_evidence_id,
        "proposal_id": row.get("proposal_id", ""),
        "claim_id": row.get("claim_id", ""),
        "source_id": row.get("source_id", ""),
        "evidence_argument": row.get("evidence_argument", ""),
        "category": row.get("category", ""),
        "source_book": row.get("source_book", ""),
        "weight_0_5": row.get("approved_weight_0_5", ""),
        "EC_MI5": row.get("EC_MI5", ""),
        "PC_MI5": row.get("PC_MI5", ""),
        "PT_MI5": row.get("PT_MI5", ""),
        "CT_MI5": row.get("CT_MI5", ""),
        "MT_MI5": row.get("MT_MI5", ""),
        "IS_MI5": row.get("IS_MI5", ""),
        "MS_MI5": row.get("MS_MI5", ""),
        "HC_MI5": row.get("HC_MI5", ""),
        "N_MI5": row.get("N_MI5", ""),
        "date": row.get("approved_date", ""),
        "notes": notes,
        "validation_status": "blocked" if errors else "ready",
        "warnings": "; ".join(warnings),
        "_errors": errors,
    }


def _read_header_row(sheet: Any, row_number: int) -> list[str]:
    values = next(sheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True), ())
    return [str(value).strip() if value is not None else "" for value in values]


def _meaningful_evidence_rows(sheet: Any, headers: list[str], start_row: int, config: dict[str, Any]) -> list[int]:
    input_columns = set(config["workbook_export"]["input_columns"].values())
    meaningful_columns = [
        index
        for index, header in enumerate(headers, start=1)
        if header in input_columns and not any(marker.lower() in header.lower() for marker in config["workbook_export"]["formula_column_markers"])
    ]
    if not meaningful_columns:
        meaningful_columns = list(range(1, len(headers) + 1))
    rows: list[int] = []
    for row_number in range(start_row, sheet.max_row + 1):
        if any(_cell_has_meaningful_value(sheet.cell(row=row_number, column=column).value) for column in meaningful_columns):
            rows.append(row_number)
    return rows


def _detect_formula_columns(sheet: Any, headers: list[str], config: dict[str, Any]) -> list[str]:
    markers = config["workbook_export"]["formula_column_markers"]
    detected = {header for header in headers if header and any(marker.lower() in header.lower() for marker in markers)}
    for column_index, header in enumerate(headers, start=1):
        if not header:
            continue
        for cells in sheet.iter_rows(min_row=config["workbook_export"]["evidence_log_header_row"] + 1, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
            value = cells[0].value
            if isinstance(value, str) and value.startswith("="):
                detected.add(header)
                break
    return sorted(detected)


def _cell_has_meaningful_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _plan_ids(sheet: Any, headers: list[str], start_row: int, last_row: int | None = None) -> dict[str, Any]:
    if "ID" not in headers:
        return {"status": "missing_id_column", "next_id": None, "warnings": ["ID column not found; planned IDs left blank."]}
    column_index = headers.index("ID") + 1
    ids: list[int] = []
    inconsistent = False
    max_row = last_row if last_row is not None else sheet.max_row
    if max_row < start_row:
        return {"status": "numeric_ids_planned", "next_id": 1, "warnings": []}
    for cells in sheet.iter_rows(min_row=start_row, max_row=max_row, min_col=column_index, max_col=column_index, values_only=True):
        value = cells[0]
        if value is None or str(value).strip() == "":
            continue
        if isinstance(value, int) or (isinstance(value, str) and value.strip().isdigit()):
            ids.append(int(value))
        else:
            inconsistent = True
    if inconsistent:
        return {"status": "inconsistent_existing_ids", "next_id": None, "warnings": ["Existing Evidence Log IDs are text-based or inconsistent; planned IDs left blank."]}
    return {"status": "numeric_ids_planned", "next_id": (max(ids) + 1 if ids else 1), "warnings": []}


def _planned_ids_for_rows(status: str, next_id: int | None, count: int) -> list[str]:
    if status != "numeric_ids_planned" or next_id is None:
        return [""] * count
    return [str(next_id + index) for index in range(count)]


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _filter_rows(rows: list[dict[str, str]], *, proposal_id: str | None, source_id: str | None, limit: int | None) -> list[dict[str, str]]:
    filtered = []
    for row in rows:
        if proposal_id is not None and row.get("proposal_id") != proposal_id:
            continue
        if source_id is not None and row.get("source_id") != source_id:
            continue
        filtered.append(row)
    return filtered[:limit] if limit is not None else filtered


def _load_reference_ids(queue_dir: Path, config: dict[str, Any]) -> dict[str, set[str]]:
    return {
        "proposal_ids": _ids_from(queue_dir / config["queues"]["files"]["proposed_updates"], "proposal_id"),
        "claim_ids": _ids_from(queue_dir / config["queues"]["files"]["extracted_claims"], "claim_id"),
        "source_ids": _ids_from(queue_dir / config["queues"]["files"]["source_dossiers"], "source_id"),
    }


def _ids_from(path: Path, field: str) -> set[str]:
    if not path.exists():
        return set()
    return {(row.get(field) or "").strip() for row in _read_csv_rows(path) if (row.get(field) or "").strip()}


def _validate_weight(value: str, proposal_id: str, errors: list[str]) -> None:
    try:
        number = float(value)
    except ValueError:
        errors.append(f"{proposal_id}: approved_weight_0_5 must be numeric from 0 to 5.")
        return
    if number < 0 or number > 5:
        errors.append(f"{proposal_id}: approved_weight_0_5 must be between 0 and 5.")


def _validate_mi5(row: dict[str, str], config: dict[str, Any], errors: list[str]) -> None:
    allowed = set(config["allowed_values"]["mi5_labels"])
    for field in MI5_COLUMNS:
        value = (row.get(field) or "").strip()
        if value and value not in allowed:
            errors.append(f"{row.get('proposal_id')}: {field} has invalid MI5 value '{value}'.")


def _validate_date(value: str, proposal_id: str, errors: list[str]) -> None:
    if len(value.strip()) < 10:
        errors.append(f"{proposal_id}: approved_date must be ISO-like date YYYY-MM-DD.")


def _combined_notes(row: dict[str, str]) -> str:
    original = row.get("notes", "")
    trace = (
        f"Trace: proposal_id={row.get('proposal_id', '')}; claim_id={row.get('claim_id', '')}; "
        f"source_id={row.get('source_id', '')}; approved_by={row.get('approved_by', '')}; "
        f"approved_date={row.get('approved_date', '')}"
    )
    return f"Original note: {original}\n{trace}" if original else trace


def _finalize(result: dict[str, Any]) -> dict[str, Any]:
    if result["errors"]:
        result["overall_status"] = "fail"
    elif result["warnings"]:
        result["overall_status"] = "warning"
    else:
        result["overall_status"] = "pass"
    if result["errors"]:
        result["next_step_notes"] = ["Fix workbook structure or approved row validation errors before Phase 7 writing."]
    elif result["warnings"]:
        result["next_step_notes"] = ["Review warnings before using this change plan in Phase 7."]
    else:
        result["next_step_notes"] = ["Preview is ready. Phase 7 can use this plan to write to a timestamped workbook copy."]
    return result


def _bullet_list(items: list[str]) -> list[str]:
    if not items:
        return ["- None"]
    return [f"- {item}" for item in items]


def _inline(items: list[str]) -> str:
    if not items:
        return "None"
    return ", ".join(f"`{item}`" for item in items)
