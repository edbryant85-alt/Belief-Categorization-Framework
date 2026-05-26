from __future__ import annotations

import hashlib
import json
from datetime import datetime, timedelta
from pathlib import Path

import yaml
from openpyxl import Workbook

from belief_dashboard.artifacts import (
    UnknownArtifactTypeError,
    find_reports,
    find_verified_outputs,
    latest_artifact,
    list_artifact_categories,
    show_artifact,
)
from belief_dashboard.cli import main
from belief_dashboard.config import load_config


def test_list_artifacts_lists_configured_artifact_categories(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)

    result = list_artifact_categories(config, tmp_path)

    artifact_types = {row["artifact_type"] for row in result["rows"]}
    assert "export_verification" in artifact_types
    assert "output_workbooks" in artifact_types
    assert "promoted_archives" in artifact_types
    assert "rollback_archives" in artifact_types


def test_list_artifacts_format_json_returns_structured_output(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)

    exit_code = main(["list-artifacts", "--config", str(config_path), "--format", "json"])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["artifact_type"] == "artifact_categories"
    assert output["rows"]


def test_list_artifacts_save_writes_navigation_reports(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)

    exit_code = main(["list-artifacts", "--config", str(config_path), "--save"])

    output = capsys.readouterr().out
    reports_dir = tmp_path / "reports" / "artifact_navigation"
    assert exit_code == 0
    assert "Markdown report:" in output
    assert "JSON report:" in output
    assert list(reports_dir.glob("artifact_categories_*.md"))
    assert list(reports_dir.glob("artifact_categories_*.json"))


def test_latest_artifact_output_workbooks_returns_newest_workbook(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    older = tmp_path / "outputs" / "older.xlsx"
    newer = tmp_path / "outputs" / "newer.xlsx"
    _create_sample_workbook(older, config)
    _create_sample_workbook(newer, config)
    _touch(older, datetime(2026, 5, 26, 10, 0, 0))
    _touch(newer, datetime(2026, 5, 26, 11, 0, 0))

    result = latest_artifact("output_workbooks", config, tmp_path)

    assert result["path"] == str(newer)
    assert result["exists"] is True


def test_latest_artifact_export_verification_returns_newest_report(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    older = tmp_path / "reports" / "export_verification" / "export_verification_old.json"
    newer = tmp_path / "reports" / "export_verification" / "export_verification_new.json"
    _write_json(older, {"overall_status": "fail"})
    _write_json(newer, {"overall_status": "pass"})
    _touch(older, datetime(2026, 5, 26, 10, 0, 0))
    _touch(newer, datetime(2026, 5, 26, 11, 0, 0))

    result = latest_artifact("export_verification", config, tmp_path)

    assert result["path"] == str(newer)
    assert result["parsed_status"] == "pass"


def test_show_artifact_summarizes_json_reports(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    report = tmp_path / "reports" / "export_verification" / "export_verification.json"
    _write_json(report, {"overall_status": "pass", "output_workbook_path": "output.xlsx", "approved_rows_considered": 2, "warnings": [], "errors": ["x"]})

    result = show_artifact(report, config)

    assert result["artifact_kind"] == "json_report"
    assert result["status"] == "pass"
    assert result["workbook_path"] == "output.xlsx"
    assert result["errors_count"] == 1


def test_show_artifact_handles_malformed_json_gracefully(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    report = tmp_path / "reports" / "export_verification" / "bad.json"
    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text("{bad", encoding="utf-8")

    result = show_artifact(report, config)

    assert result["artifact_kind"] == "json_report"
    assert result["json_valid"] is False
    assert result["parse_error"]


def test_show_artifact_summarizes_markdown_without_dumping_huge_files(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    report = tmp_path / "reports" / "workbook_recovery" / "report.md"
    report.parent.mkdir(parents=True, exist_ok=True)
    report.write_text("# Heading\n\n" + "\n".join(f"line {i}" for i in range(50)), encoding="utf-8")

    result = show_artifact(report, config)

    assert result["artifact_kind"] == "markdown_report"
    assert result["preview_lines"][0] == "# Heading"
    assert len(result["preview_lines"]) == 8


def test_show_artifact_summarizes_workbook_without_modifying_it(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    workbook = tmp_path / "outputs" / "output.xlsx"
    _create_sample_workbook(workbook, config)
    before = _sha256(workbook)

    result = show_artifact(workbook, config)

    assert result["artifact_kind"] == "workbook"
    assert result["inspection_status"] == "pass"
    assert _sha256(workbook) == before


def test_find_report_status_pass_filters_correctly(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    _write_json(tmp_path / "reports" / "export_verification" / "pass.json", {"overall_status": "pass"})
    _write_json(tmp_path / "reports" / "export_verification" / "fail.json", {"overall_status": "fail"})

    result = find_reports("export_verification", config, tmp_path, status="pass")

    assert result["count"] == 1
    assert result["rows"][0]["path"].endswith("pass.json")


def test_find_report_contains_filters_correctly(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    _write_json(tmp_path / "reports" / "reviews" / "one.json", {"overall_status": "pass", "proposal_id": "PROP0001"})
    _write_json(tmp_path / "reports" / "reviews" / "two.json", {"overall_status": "pass", "proposal_id": "PROP9999"})

    result = find_reports("reviews", config, tmp_path, contains="PROP0001")

    assert result["count"] == 1
    assert result["rows"][0]["path"].endswith("one.json")


def test_find_verified_output_lists_passing_verification_reports(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    workbook = tmp_path / "outputs" / "output.xlsx"
    _create_sample_workbook(workbook, config)
    _write_json(
        tmp_path / "reports" / "export_verification" / "verification.json",
        {"overall_status": "pass", "output_workbook_path": str(workbook), "verification_timestamp": "2026-05-26T12:00:00"},
    )

    result = find_verified_outputs(config, tmp_path, status="pass")

    assert result["count"] == 1
    assert result["rows"][0]["output_workbook_path"] == str(workbook)
    assert result["rows"][0]["output_workbook_exists"] is True


def test_find_verified_output_latest_returns_only_latest_match(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)
    older = tmp_path / "outputs" / "older.xlsx"
    newer = tmp_path / "outputs" / "newer.xlsx"
    _create_sample_workbook(older, config)
    _create_sample_workbook(newer, config)
    _write_json(
        tmp_path / "reports" / "export_verification" / "older.json",
        {"overall_status": "pass", "output_workbook_path": str(older), "verification_timestamp": "2026-05-26T10:00:00"},
    )
    _write_json(
        tmp_path / "reports" / "export_verification" / "newer.json",
        {"overall_status": "pass", "output_workbook_path": str(newer), "verification_timestamp": "2026-05-26T11:00:00"},
    )

    result = find_verified_outputs(config, tmp_path, status="pass", latest=True)

    assert result["count"] == 1
    assert result["rows"][0]["output_workbook_path"] == str(newer)


def test_unknown_artifact_type_fails_clearly(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)

    try:
        latest_artifact("not_a_type", config, tmp_path)
    except UnknownArtifactTypeError as exc:
        assert "Unknown artifact type" in str(exc)
    else:
        raise AssertionError("Expected UnknownArtifactTypeError")


def test_missing_artifact_path_fails_clearly(tmp_path: Path) -> None:
    config = _artifact_config(tmp_path)

    try:
        show_artifact(tmp_path / "missing.json", config)
    except FileNotFoundError as exc:
        assert "Artifact not found" in str(exc)
    else:
        raise AssertionError("Expected FileNotFoundError")


def _artifact_config(tmp_path: Path) -> dict:
    config = load_config("config.yaml")
    config["artifact_navigation"] = {
        "reports_dir": str(tmp_path / "reports" / "artifact_navigation"),
        "reports": {
            "workbook_inspection": str(tmp_path / "reports" / "workbook_inspection"),
            "queue_validation": str(tmp_path / "reports" / "queue_validation"),
            "prompt_packets": str(tmp_path / "reports" / "prompt_packets"),
            "manual_imports": str(tmp_path / "reports" / "manual_imports"),
            "reviews": str(tmp_path / "reports" / "reviews"),
            "workbook_export_preview": str(tmp_path / "reports" / "workbook_export_preview"),
            "workbook_exports": str(tmp_path / "reports" / "workbook_exports"),
            "export_verification": str(tmp_path / "reports" / "export_verification"),
            "workbook_promotion": str(tmp_path / "reports" / "workbook_promotion"),
            "workbook_recovery": str(tmp_path / "reports" / "workbook_recovery"),
        },
        "workbooks": {
            "main": str(tmp_path / "workbooks" / "main.xlsx"),
            "outputs": str(tmp_path / "outputs"),
            "backups": str(tmp_path / "backups"),
            "promoted_archives": str(tmp_path / "backups" / "promoted_archives"),
            "rollback_archives": str(tmp_path / "backups" / "rollback_archives"),
        },
        "default_limit": 10,
    }
    Path(config["artifact_navigation"]["reports_dir"]).mkdir(parents=True, exist_ok=True)
    for directory in config["artifact_navigation"]["reports"].values():
        Path(directory).mkdir(parents=True, exist_ok=True)
    for key, value in config["artifact_navigation"]["workbooks"].items():
        path = Path(value)
        if key == "main":
            path.parent.mkdir(parents=True, exist_ok=True)
        else:
            path.mkdir(parents=True, exist_ok=True)
    return config


def _write_config(tmp_path: Path) -> Path:
    config = _artifact_config(tmp_path)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config), encoding="utf-8")
    return config_path


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def _create_sample_workbook(path: Path, config: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]
    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)
    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)
    evidence.cell(row=header_row + 1, column=1, value="E-001")
    workbook.save(path)


def _touch(path: Path, value: datetime) -> None:
    timestamp = value.timestamp()
    path.touch()
    import os

    os.utime(path, (timestamp, timestamp))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
