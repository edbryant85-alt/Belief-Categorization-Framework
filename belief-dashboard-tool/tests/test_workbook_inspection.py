from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook

from belief_dashboard.cli import main
from belief_dashboard.config import load_config
from belief_dashboard.workbook import inspect_workbook, write_reports


def test_expected_sheet_names_are_read_from_config() -> None:
    config = load_config("config.yaml")

    assert config["workbook"]["expected_sheets"] == [
        "Dashboard",
        "MI5 Scale",
        "Hypotheses",
        "Evidence Log",
        "Decision Nodes",
        "Reading Tracker",
        "Instructions",
    ]


def test_workbook_inspection_returns_structured_result(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    workbook_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(workbook_path, config)

    result = inspect_workbook(
        workbook_path,
        config,
        inspected_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    assert result["workbook_file_exists"] is True
    assert result["overall_status"] == "pass"
    assert result["expected_sheets"]["missing"] == []
    assert result["evidence_log"]["required_columns"]["missing"] == []
    assert result["evidence_log"]["hypothesis_mi5_columns"]["missing"] == []
    assert result["evidence_log"]["populated_evidence_rows"] == 2


def test_missing_workbook_path_produces_clear_error(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    missing_path = tmp_path / "missing.xlsx"

    result = inspect_workbook(missing_path, config)

    assert result["workbook_file_exists"] is False
    assert result["overall_status"] == "fail"
    assert "Workbook file was not found" in result["next_step_notes"][0]


def test_markdown_and_json_reports_can_be_written(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    workbook_path = tmp_path / "sample.xlsx"
    reports_dir = tmp_path / "reports"
    _create_sample_workbook(workbook_path, config)
    result = inspect_workbook(
        workbook_path,
        config,
        inspected_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    markdown_path, json_path = write_reports(
        result,
        reports_dir,
        written_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    assert markdown_path.exists()
    assert json_path.exists()
    assert "Workbook Inspection Report" in markdown_path.read_text(encoding="utf-8")
    loaded = json.loads(json_path.read_text(encoding="utf-8"))
    assert loaded["overall_status"] == "pass"


def test_cli_inspect_workbook_with_explicit_path(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    workbook_path = tmp_path / "sample.xlsx"
    reports_dir = tmp_path / "reports"
    config["paths"]["reports_dir"] = str(reports_dir)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config), encoding="utf-8")
    _create_sample_workbook(workbook_path, config)

    exit_code = main(
        [
            "inspect-workbook",
            "--config",
            str(config_path),
            "--workbook",
            str(workbook_path),
        ]
    )

    assert exit_code == 0
    assert list(reports_dir.glob("workbook_inspection_*.md"))
    assert list(reports_dir.glob("workbook_inspection_*.json"))


def _create_sample_workbook(path: Path, config: dict) -> None:
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]

    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)

    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    evidence["A1"] = "Evidence Log"
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)

    evidence.cell(row=header_row + 1, column=1, value="E-001")
    evidence.cell(row=header_row + 1, column=3, value="Placeholder evidence one")
    evidence.cell(row=header_row + 2, column=1, value="E-002")
    evidence.cell(row=header_row + 2, column=3, value="Placeholder evidence two")

    workbook.save(path)
