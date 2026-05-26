from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import Workbook

from belief_dashboard.config import load_config
from belief_dashboard.queues import init_queues
from belief_dashboard.workbook_promotion import (
    promote_output_workbook,
    write_workbook_promotion_reports,
)


PROMOTED_AT = datetime.now().replace(microsecond=0) + timedelta(days=1)


def test_dry_run_promotion_writes_no_files_and_does_not_replace_main_workbook(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    original_main_hash = _sha256(fixture["main"])

    result = _promote(fixture, dry_run=True)

    assert result["overall_status"] == "pass"
    assert result["main_workbook_replaced"] is False
    assert _sha256(fixture["main"]) == original_main_hash
    assert not list(fixture["archive_dir"].glob("*.xlsx"))
    assert not list(fixture["reports_dir"].glob("*"))
    assert _read_rows(fixture["queue_dir"] / "change_log.csv") == []


def test_promotion_fails_when_candidate_workbook_is_missing(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    fixture["candidate"].unlink()

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert any("Candidate output workbook not found" in error for error in result["errors"])
    assert result["main_workbook_replaced"] is False


def test_promotion_fails_when_main_workbook_is_missing(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    fixture["main"].unlink()

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert any("Main workbook not found" in error for error in result["errors"])


def test_promotion_fails_when_verification_report_is_missing(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    fixture["verification_report"].unlink()

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert any("Verification report not found" in error for error in result["errors"])


def test_promotion_fails_when_verification_status_is_not_acceptable(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path, verification_status="warning")

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert any("is not accepted" in error for error in result["errors"])


def test_promotion_fails_when_verification_report_refers_to_different_workbook(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path, report_workbook_path=tmp_path / "other.xlsx")

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert result["verification_report_matched_candidate"] is False
    assert any("does not refer" in error for error in result["errors"])


def test_promotion_creates_archive_replaces_main_and_preserves_output_workbook(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    original_main_hash = _sha256(fixture["main"])
    candidate_hash = _sha256(fixture["candidate"])

    result = _promote(fixture)

    assert result["overall_status"] == "pass"
    assert result["main_workbook_replaced"] is True
    archive_path = Path(result["archive_path"])
    assert archive_path.exists()
    assert _sha256(archive_path) == original_main_hash
    assert _sha256(fixture["main"]) == candidate_hash
    assert fixture["candidate"].exists()
    assert _sha256(fixture["candidate"]) == candidate_hash


def test_promotion_writes_markdown_and_json_reports(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    result = _promote(fixture)

    markdown_path, json_path = write_workbook_promotion_reports(
        result,
        fixture["reports_dir"],
        written_at=PROMOTED_AT,
    )

    assert markdown_path.exists()
    assert json_path.exists()
    assert "Workbook Promotion Report" in markdown_path.read_text(encoding="utf-8")
    assert json.loads(json_path.read_text(encoding="utf-8"))["overall_status"] == "pass"


def test_promotion_writes_to_change_log(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)

    result = _promote(fixture)

    rows = _read_rows(fixture["queue_dir"] / "change_log.csv")
    assert result["change_log_updated"] is True
    assert rows[-1]["operation"] == "promote_output_workbook"
    assert rows[-1]["input_file"] == str(fixture["candidate"])
    assert rows[-1]["output_file"] == str(fixture["main"])
    assert rows[-1]["status"] == "pass"


def test_promotion_refuses_to_overwrite_existing_archive(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    archive_path = fixture["archive_dir"] / f"{fixture['main'].stem}_pre_promotion_{PROMOTED_AT.strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    archive_path.parent.mkdir(parents=True)
    archive_path.write_text("existing", encoding="utf-8")

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert archive_path.read_text(encoding="utf-8") == "existing"
    assert any("Archive path already exists" in error for error in result["errors"])


def test_promotion_refuses_to_overwrite_existing_report_files(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    fixture["reports_dir"].mkdir(parents=True)
    stamp = PROMOTED_AT.strftime("%Y-%m-%d_%H%M%S")
    (fixture["reports_dir"] / f"workbook_promotion_{stamp}.json").write_text("existing", encoding="utf-8")

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert result["main_workbook_replaced"] is False
    assert any("Promotion report path already exists" in error for error in result["errors"])


def test_promotion_performs_basic_workbook_inspection(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path, invalid_candidate=True)

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert result["basic_workbook_inspection_passed"] is False
    assert result["basic_workbook_inspection_status"] == "fail"


def test_promotion_fails_when_candidate_changed_after_verification(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path, verification_timestamp=datetime(2000, 1, 1, 0, 0, 0))

    result = _promote(fixture)

    assert result["overall_status"] == "fail"
    assert any("changed after verification" in error for error in result["errors"])


def test_write_reports_refuses_to_overwrite_existing_files(tmp_path: Path) -> None:
    fixture = _setup_promotion_fixture(tmp_path)
    result = _promote(fixture)
    write_workbook_promotion_reports(result, fixture["reports_dir"], written_at=PROMOTED_AT)

    with pytest.raises(FileExistsError):
        write_workbook_promotion_reports(result, fixture["reports_dir"], written_at=PROMOTED_AT)


def _setup_promotion_fixture(
    tmp_path: Path,
    *,
    verification_status: str = "pass",
    report_workbook_path: Path | None = None,
    verification_timestamp: datetime | None = None,
    invalid_candidate: bool = False,
) -> dict[str, Path | dict]:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    archive_dir = tmp_path / "archives"
    reports_dir = tmp_path / "reports"
    main = tmp_path / "workbooks" / "main.xlsx"
    candidate = tmp_path / "outputs" / "output.xlsx"
    verification_report = tmp_path / "verification" / "export_verification.json"
    init_queues(queue_dir, config)
    _create_sample_workbook(main, config, marker="old-main")
    if invalid_candidate:
        _create_invalid_workbook(candidate)
    else:
        _create_sample_workbook(candidate, config, marker="verified-output")
    _write_verification_report(
        verification_report,
        workbook_path=report_workbook_path or candidate,
        status=verification_status,
        verified_at=verification_timestamp or PROMOTED_AT,
    )
    return {
        "config": config,
        "queue_dir": queue_dir,
        "archive_dir": archive_dir,
        "reports_dir": reports_dir,
        "main": main,
        "candidate": candidate,
        "verification_report": verification_report,
    }


def _promote(fixture: dict[str, Path | dict], *, dry_run: bool = False) -> dict:
    return promote_output_workbook(
        fixture["candidate"],
        fixture["verification_report"],
        fixture["main"],
        fixture["config"],
        archive_dir=fixture["archive_dir"],
        reports_dir=fixture["reports_dir"],
        queue_dir=fixture["queue_dir"],
        dry_run=dry_run,
        promoted_at=PROMOTED_AT,
    )


def _create_sample_workbook(path: Path, config: dict, *, marker: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]
    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)
    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)
    evidence.cell(row=header_row + 1, column=1, value=marker)
    evidence.cell(row=header_row + 1, column=3, value=f"{marker} evidence")
    workbook.save(path)


def _create_invalid_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Not Evidence Log"
    workbook.save(path)


def _write_verification_report(path: Path, *, workbook_path: Path, status: str, verified_at: datetime) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(
            {
                "output_workbook_path": str(workbook_path),
                "verification_timestamp": verified_at.replace(microsecond=0).isoformat(),
                "overall_status": status,
            },
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))
