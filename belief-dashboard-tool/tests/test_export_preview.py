from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from belief_dashboard.config import load_config
from belief_dashboard.export_preview import preview_workbook_export, write_export_preview_artifacts
from belief_dashboard.queues import init_queues
from belief_dashboard.schemas import QUEUE_SCHEMAS


def test_preview_export_reads_workbook_structure_without_modifying_it(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path)
    before = workbook_path.read_bytes()

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "pass"
    assert workbook_path.read_bytes() == before


def test_preview_export_identifies_evidence_log_input_columns(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path)

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert "Evidence / Argument" in result["workbook_input_columns"]["found"]
    assert result["workbook_input_columns"]["missing"] == []


def test_preview_export_identifies_formula_columns(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path)

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert "EC Numeric" in result["formula_driven_columns_detected"]
    assert "EC Factor" in result["formula_driven_columns_detected"]
    assert "EC Log Factor" in result["formula_driven_columns_detected"]


def test_valid_approved_rows_produce_change_plan(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path)

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["rows_ready_for_export"] == 1
    planned = result["planned_rows"][0]
    assert planned["planned_workbook_row"] == 5
    assert planned["planned_evidence_id"] == "2"
    assert planned["evidence_argument"] == "Approved evidence."
    assert planned["date"] == "2026-05-25"
    assert "Trace: proposal_id=PROP0001" in planned["notes"]


def test_missing_required_approved_fields_fail_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"evidence_argument": ""})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("evidence_argument is required" in error for error in result["errors"])


def test_invalid_mi5_labels_fail_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"EC_MI5": "Certain-ish"})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("EC_MI5 has invalid MI5 value" in error for error in result["errors"])


def test_out_of_range_approved_weights_fail_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"approved_weight_0_5": "6"})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("approved_weight_0_5 must be between 0 and 5" in error for error in result["errors"])


def test_missing_referenced_source_id_fails_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"source_id": "SRC9999"})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("source_id not found" in error for error in result["errors"])


def test_missing_referenced_claim_id_fails_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"claim_id": "C999"})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("claim_id not found" in error for error in result["errors"])


def test_missing_referenced_proposal_id_fails_validation(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, approved_overrides={"proposal_id": "PROP9999"})

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("proposal_id not found" in error for error in result["errors"])


def test_numeric_existing_ids_produce_planned_next_ids(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, existing_id="41")

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["id_planning_status"] == "numeric_ids_planned"
    assert result["planned_rows"][0]["planned_evidence_id"] == "42"


def test_inconsistent_existing_ids_produce_warnings_and_blank_planned_ids(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path, existing_id="E-001")

    result = preview_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "warning"
    assert result["id_planning_status"] == "inconsistent_existing_ids"
    assert result["planned_rows"][0]["planned_evidence_id"] == ""


def test_markdown_json_and_csv_change_plan_files_are_written(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_preview_fixture(tmp_path)
    result = preview_workbook_export(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        previewed_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    markdown_path, json_path, csv_path = write_export_preview_artifacts(
        result,
        tmp_path / "reports",
        written_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    assert markdown_path.name == "workbook_export_preview_2026-05-25_153000.md"
    assert json_path.name == "workbook_export_preview_2026-05-25_153000.json"
    assert csv_path.name == "workbook_export_change_plan_2026-05-25_153000.csv"
    assert "Workbook Export Preview" in markdown_path.read_text(encoding="utf-8")
    assert json.loads(json_path.read_text(encoding="utf-8"))["overall_status"] == "pass"
    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    assert rows[0]["proposal_id"] == "PROP0001"


def _setup_preview_fixture(
    tmp_path: Path,
    *,
    approved_overrides: dict[str, str] | None = None,
    existing_id: str = "1",
) -> tuple[dict, Path, Path]:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    init_queues(queue_dir, config)
    workbook_path = tmp_path / "workbook.xlsx"
    _create_workbook(workbook_path, existing_id=existing_id)
    _append_queue_row(queue_dir / "source_dossiers.csv", "source_dossiers", {"source_id": "SRC0001"})
    _append_queue_row(queue_dir / "extracted_claims.csv", "extracted_claims", {"claim_id": "C001", "source_id": "SRC0001", "claim_text": "Claim."})
    _append_queue_row(queue_dir / "proposed_updates.csv", "proposed_updates", {"proposal_id": "PROP0001", "claim_id": "C001", "source_id": "SRC0001"})
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row(approved_overrides))
    return config, queue_dir, workbook_path


def _create_workbook(path: Path, *, existing_id: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence Log"
    sheet["A1"] = "Evidence Log"
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "EC Numeric",
        "EC Factor",
        "EC Log Factor",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index, value=header)
    sheet.cell(row=4, column=1, value=existing_id)
    sheet.cell(row=4, column=3, value="Existing evidence.")
    sheet.cell(row=4, column=16, value="=1")
    workbook.save(path)


def _approved_row(overrides: dict[str, str] | None = None) -> dict[str, str]:
    row = {
        "proposal_id": "PROP0001",
        "claim_id": "C001",
        "source_id": "SRC0001",
        "evidence_argument": "Approved evidence.",
        "category": "Example",
        "source_book": "Example Source",
        "approved_weight_0_5": "3",
        "EC_MI5": "Likely / probable",
        "PC_MI5": "Roughly even chance",
        "notes": "Approved note.",
        "approved_by": "Eric",
        "approved_date": "2026-05-25",
    }
    if overrides:
        row.update(overrides)
    return row


def _append_queue_row(path: Path, queue_name: str, row: dict[str, str]) -> None:
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUEUE_SCHEMAS[queue_name])
        writer.writerow({header: row.get(header, "") for header in QUEUE_SCHEMAS[queue_name]})
