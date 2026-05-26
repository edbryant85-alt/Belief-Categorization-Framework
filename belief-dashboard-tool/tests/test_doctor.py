from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook

from belief_dashboard.cli import main
from belief_dashboard.config import load_config
from belief_dashboard.doctor import build_doctor_explanation, build_doctor_report
from belief_dashboard.queues import init_queues
from belief_dashboard.schemas import QUEUE_SCHEMAS


def test_doctor_works_in_general_mode(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    init_queues(tmp_path / "queues", config)

    result = build_doctor_report(config, tmp_path, mode="general")

    assert result["operation"] == "doctor"
    assert result["mode"] == "general"
    assert result["overall_status"] in {"pass", "warning"}
    assert result["next_safest_commands"]


def test_doctor_format_json_returns_structured_output(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--format", "json"])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["operation"] == "doctor"
    assert output["summary_counts_by_severity"]
    assert isinstance(output["findings"], list)


def test_doctor_save_writes_markdown_and_json_reports(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--save"])

    output = capsys.readouterr().out
    reports_dir = tmp_path / "reports" / "doctor"
    assert exit_code == 0
    assert "Markdown report:" in output
    assert "JSON report:" in output
    assert list(reports_dir.glob("doctor_GENERAL_*.md"))
    assert list(reports_dir.glob("doctor_GENERAL_*.json"))


def test_missing_main_workbook_produces_blocker_finding(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    init_queues(tmp_path / "queues", config)

    result = build_doctor_report(config, tmp_path, mode="general")

    finding = _finding(result, "main_workbook_missing")
    assert finding["severity"] == "blocker"
    assert "inspect-workbook" in finding["safe_repair_command"]
    assert finding["documentation_reference"]


def test_missing_queues_produce_blocker_with_init_guidance(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)

    result = build_doctor_report(config, tmp_path, mode="general")

    finding = _finding(result, "queues_missing")
    assert finding["severity"] == "blocker"
    assert "init-queues" in finding["safe_repair_command"]
    assert "validate-queues" in finding["safe_repair_command"]


def test_queue_validation_failure_produces_error_finding(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    init_queues(tmp_path / "queues", config)
    _append_row(
        tmp_path / "queues" / "proposed_updates.csv",
        QUEUE_SCHEMAS["proposed_updates"],
        {"proposal_id": "PROP0001", "EC_MI5": "Certain-ish", "review_status": "proposed"},
    )

    result = build_doctor_report(config, tmp_path, mode="general")

    finding = _finding(result, "queue_validation_failed")
    assert finding["severity"] == "error"
    assert finding["safe_repair_command"] == "python -m belief_dashboard.cli validate-queues"
    assert "reports/queue_validation" in finding["documentation_reference"]


def test_doctor_before_export_fails_when_workbook_or_queues_are_not_ready(tmp_path: Path) -> None:
    config_path = _write_config(tmp_path)

    exit_code = main(["doctor", "--config", str(config_path), "--mode", "before-export"])

    assert exit_code == 1


def test_doctor_before_verification_fails_when_no_output_workbook_exists(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path)

    exit_code = main(["doctor", "--config", str(config_path), "--mode", "before-verification"])

    assert exit_code == 1


def test_doctor_before_promotion_fails_when_no_passing_verification_exists(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path)
    _write_file(tmp_path / "outputs" / "output.xlsx", b"output")

    exit_code = main(["doctor", "--config", str(config_path), "--mode", "before-promotion"])

    assert exit_code == 1


def test_doctor_before_rollback_fails_when_no_promoted_archive_exists(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path)

    exit_code = main(["doctor", "--config", str(config_path), "--mode", "before-rollback"])

    assert exit_code == 1


def test_findings_include_required_guidance_fields(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)

    result = build_doctor_report(config, tmp_path, mode="general")

    finding = result["findings"][0]
    assert finding["severity"]
    assert finding["plain_language_explanation"]
    assert finding["safe_repair_command"]
    assert finding["documentation_reference"]
    assert finding["can_auto_fix"] is False


def test_doctor_does_not_modify_workbooks(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path)
    config = load_config(config_path)
    workbook = Path(config["workbook"]["default_path"])
    before = _sha256(workbook)

    assert main(["doctor", "--config", str(config_path), "--mode", "before-export"]) in {0, 1}

    assert _sha256(workbook) == before


def test_doctor_does_not_modify_queue_csv_files(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path, approved_rows=1)
    config = load_config(config_path)
    queue_dir = Path(config["queues"]["base_dir"])
    before = {path.name: _sha256(path) for path in queue_dir.glob("*.csv")}

    assert main(["doctor", "--config", str(config_path), "--mode", "before-export"]) in {0, 1}

    after = {path.name: _sha256(path) for path in queue_dir.glob("*.csv")}
    assert after == before


def test_doctor_explain_existing_finding_prints_detail(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--explain", "MAIN_WORKBOOK_MISSING"])

    output = capsys.readouterr().out
    assert exit_code == 0
    assert "Finding: main_workbook_missing" in output
    assert "Likely causes:" in output
    assert "Safest next steps:" in output
    assert "Do not:" in output


def test_doctor_before_promotion_explain_no_passing_verification(tmp_path: Path, capsys) -> None:
    config_path = _ready_config_path(tmp_path)

    exit_code = main(
        [
            "doctor",
            "--config",
            str(config_path),
            "--mode",
            "before-promotion",
            "--explain",
            "NO_PASSING_VERIFICATION",
        ]
    )

    output = capsys.readouterr().out
    assert exit_code == 0
    assert "Finding: no_passing_verification" in output
    assert "find-verified-output" in output
    assert "verify-workbook-export" in output


def test_doctor_explain_unknown_id_lists_detected_findings(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--explain", "UNKNOWN_ID"])

    output = capsys.readouterr().out
    assert exit_code == 1
    assert "Finding not currently detected: UNKNOWN_ID" in output
    assert "Mode checked: general" in output
    assert "main_workbook_missing" in output
    assert "python -m belief_dashboard.cli doctor --mode general" in output


def test_doctor_explain_format_json_returns_structured_explanation(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--explain", "main-workbook-missing", "--format", "json"])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["operation"] == "doctor_explain"
    assert output["finding_id"] == "main_workbook_missing"
    assert output["likely_causes"]
    assert output["safest_next_steps"]
    assert output["safe_repair_commands"]
    assert output["documentation_references"]
    assert output["what_not_to_do"]


def test_doctor_explain_save_writes_markdown_and_json_reports(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    init_queues(tmp_path / "queues", config)

    exit_code = main(["doctor", "--config", str(config_path), "--explain", "MAIN_WORKBOOK_MISSING", "--save"])

    output = capsys.readouterr().out
    reports_dir = tmp_path / "reports" / "doctor"
    assert exit_code == 0
    assert "Markdown report:" in output
    assert "JSON report:" in output
    assert list(reports_dir.glob("doctor_explain_MAIN_WORKBOOK_MISSING_*.md"))
    assert list(reports_dir.glob("doctor_explain_MAIN_WORKBOOK_MISSING_*.json"))


def test_doctor_explanation_model_includes_guidance_sections(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    init_queues(tmp_path / "queues", config)

    explanation = build_doctor_explanation(config, tmp_path, "main_workbook_missing")

    assert explanation["status"] == "detected"
    assert explanation["likely_causes"]
    assert explanation["safest_next_steps"]
    assert explanation["safe_repair_commands"]
    assert explanation["documentation_references"]
    assert explanation["what_not_to_do"]
    assert explanation["auto_fix_available"] is False


def test_doctor_explain_missing_queues_includes_init_guidance(tmp_path: Path) -> None:
    config = _doctor_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)

    explanation = build_doctor_explanation(config, tmp_path, "MISSING_QUEUES")

    assert explanation["status"] == "detected"
    assert explanation["finding_id"] == "queues_missing"
    assert "python -m belief_dashboard.cli init-queues" in explanation["safe_repair_commands"]
    assert any("Do not manually invent queue headers" in item for item in explanation["what_not_to_do"])


def test_doctor_explain_does_not_modify_workbooks(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path)
    config = load_config(config_path)
    workbook = Path(config["workbook"]["default_path"])
    before = _sha256(workbook)

    assert main(["doctor", "--config", str(config_path), "--explain", "no_verified_output_general"]) == 0

    assert _sha256(workbook) == before


def test_doctor_explain_does_not_modify_queue_csv_files(tmp_path: Path) -> None:
    config_path = _ready_config_path(tmp_path, approved_rows=1)
    config = load_config(config_path)
    queue_dir = Path(config["queues"]["base_dir"])
    before = {path.name: _sha256(path) for path in queue_dir.glob("*.csv")}

    assert main(["doctor", "--config", str(config_path), "--explain", "no_verified_output_general"]) == 0

    after = {path.name: _sha256(path) for path in queue_dir.glob("*.csv")}
    assert after == before


def _ready_config_path(tmp_path: Path, *, approved_rows: int = 0) -> Path:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    init_queues(tmp_path / "queues", config)
    if approved_rows:
        _append_row(
            tmp_path / "queues" / "approved_updates.csv",
            QUEUE_SCHEMAS["approved_updates"],
            {
                "proposal_id": "PROP0001",
                "claim_id": "CLM0001",
                "source_id": "SRC0001",
                "evidence_argument": "Evidence",
                "category": "Test",
                "source_book": "Source",
            },
        )
    return config_path


def _doctor_config(tmp_path: Path) -> dict:
    config = load_config("config.yaml")
    config["workbook"]["default_path"] = str(tmp_path / "workbooks" / "main.xlsx")
    config["queues"]["base_dir"] = str(tmp_path / "queues")
    config["paths"]["sample_dir"] = str(tmp_path / "sample")
    config["product_readiness"]["sample_demo_dir"] = str(tmp_path / "sample" / "end_to_end_demo")
    config["doctor"]["reports_dir"] = str(tmp_path / "reports" / "doctor")
    config["artifact_navigation"] = {
        "reports_dir": str(tmp_path / "reports" / "artifact_navigation"),
        "reports": {
            "workbook_inspection": str(tmp_path / "reports" / "workbook_inspection"),
            "queue_validation": str(tmp_path / "reports" / "queue_validation"),
            "prompt_packets": str(tmp_path / "reports" / "prompt_packets"),
            "manual_imports": str(tmp_path / "reports" / "manual_imports"),
            "reviews": str(tmp_path / "reports" / "reviews"),
            "workbook_export_preview": str(tmp_path / "reports" / "workbook_export_preview"),
            "workbook_exports": str(tmp_path / "reports" / "workbook_exports"),
            "export_verification": str(tmp_path / "reports" / "export_verification"),
            "workbook_promotion": str(tmp_path / "reports" / "workbook_promotion"),
            "workbook_recovery": str(tmp_path / "reports" / "workbook_recovery"),
        },
        "workbooks": {
            "main": str(tmp_path / "workbooks" / "main.xlsx"),
            "outputs": str(tmp_path / "outputs"),
            "backups": str(tmp_path / "backups"),
            "promoted_archives": str(tmp_path / "backups" / "promoted_archives"),
            "rollback_archives": str(tmp_path / "backups" / "rollback_archives"),
        },
        "default_limit": 10,
    }
    config["command_composition"] = {
        "reports_dir": str(tmp_path / "reports" / "command_guides"),
        "default_include_dry_run_first": True,
        "quote_paths": True,
    }
    config["operator_preflight"] = {
        "reports_dir": str(tmp_path / "reports" / "operator_preflight"),
        "default_mode": "general",
    }
    config["workbook_export"]["outputs_dir"] = str(tmp_path / "outputs")
    config["export_verification"]["outputs_dir"] = str(tmp_path / "outputs")
    config["workbook_promotion"]["archive_dir"] = str(tmp_path / "backups" / "promoted_archives")
    config["workbook_recovery"]["archive_dir"] = str(tmp_path / "backups" / "promoted_archives")

    for directory in config["artifact_navigation"]["reports"].values():
        Path(directory).mkdir(parents=True, exist_ok=True)
    for directory in [
        config["artifact_navigation"]["reports_dir"],
        config["artifact_navigation"]["workbooks"]["outputs"],
        config["artifact_navigation"]["workbooks"]["promoted_archives"],
        config["artifact_navigation"]["workbooks"]["rollback_archives"],
        config["command_composition"]["reports_dir"],
        config["doctor"]["reports_dir"],
        config["operator_preflight"]["reports_dir"],
    ]:
        Path(directory).mkdir(parents=True, exist_ok=True)
    _create_demo_assets(Path(config["product_readiness"]["sample_demo_dir"]))
    Path(config["workbook"]["default_path"]).parent.mkdir(parents=True, exist_ok=True)
    return config


def _write_config(tmp_path: Path) -> Path:
    config = _doctor_config(tmp_path)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config), encoding="utf-8")
    return config_path


def _create_sample_workbook(path: Path, config: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]
    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)
    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)
    workbook.save(path)


def _create_demo_assets(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for filename in [
        "sample_source.md",
        "extracted_claims.csv",
        "criteria_matrix.csv",
        "proposed_updates.csv",
        "README.md",
    ]:
        (path / filename).write_text("sample\n", encoding="utf-8")
    workbook = Workbook()
    workbook.save(path / "demo_workbook.xlsx")


def _append_row(path: Path, headers: list[str], values: dict[str, str]) -> None:
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        row = {header: "" for header in headers}
        row.update(values)
        writer.writerow(row)


def _write_file(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def _finding(result: dict, finding_id: str) -> dict:
    return next(finding for finding in result["findings"] if finding["id"] == finding_id)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
