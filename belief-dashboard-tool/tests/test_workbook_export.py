from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from belief_dashboard.config import load_config
from belief_dashboard.queues import init_queues
from belief_dashboard.schemas import QUEUE_SCHEMAS
from belief_dashboard.workbook_export import apply_approved_to_workbook, write_workbook_export_report


def test_apply_approved_to_workbook_dry_run_writes_no_workbook_files(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
        dry_run=True,
    )

    assert result["overall_status"] == "pass"
    assert not (tmp_path / "backups").exists()
    assert not (tmp_path / "outputs").exists()


def test_real_export_creates_backup_workbook(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    assert Path(result["backup_workbook_path"]).exists()


def test_real_export_creates_output_workbook(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    assert Path(result["output_workbook_path"]).exists()


def test_original_workbook_remains_unchanged(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path)
    before = workbook_path.read_bytes()

    apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    assert workbook_path.read_bytes() == before


def test_approved_rows_are_appended_to_output_evidence_log(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    sheet = _output_sheet(result)
    assert sheet.cell(row=5, column=3).value == "Approved evidence."


def test_input_columns_receive_correct_mapped_values(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    sheet = _output_sheet(result)
    assert sheet.cell(row=5, column=1).value == "2"
    assert sheet.cell(row=5, column=2).value == "2026-05-25"
    assert sheet.cell(row=5, column=4).value == "Example"
    assert sheet.cell(row=5, column=6).value == "3"
    assert sheet.cell(row=5, column=7).value == "Likely / probable"
    assert "Trace: proposal_id=PROP0001" in sheet.cell(row=5, column=19).value


def test_formula_columns_are_copied_down_where_safe(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    sheet = _output_sheet(result)
    assert sheet.cell(row=5, column=16).value == "=A5"
    assert "EC Numeric" in result["formula_driven_columns_copied"]


def test_export_appends_after_last_meaningful_row_not_stale_excel_range(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, stale_blank_row=251)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    sheet = _output_sheet(result)
    assert result["first_appended_row"] == 5
    assert sheet.cell(row=5, column=3).value == "Approved evidence."


def test_formula_copy_uses_last_actual_formula_row_before_blank_gap(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, formula_row=68, stale_blank_row=251)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    sheet = _output_sheet(result)
    assert result["first_appended_row"] == 69
    assert sheet.cell(row=69, column=16).value == "=A69"
    assert "EC Numeric" in result["formula_driven_columns_copied"]


def test_formula_columns_are_not_overwritten_with_queue_values(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    sheet = _output_sheet(result)
    assert sheet.cell(row=5, column=16).value != "Likely / probable"


def test_invalid_approved_rows_prevent_export(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, approved_overrides={"approved_weight_0_5": "9"})

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    assert result["overall_status"] == "fail"
    assert not (tmp_path / "outputs").exists()


def test_missing_required_workbook_columns_prevent_export(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, omit_notes=True)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    assert result["overall_status"] == "fail"
    assert not (tmp_path / "outputs").exists()


def test_inconsistent_existing_ids_leave_planned_ids_blank_but_allow_export_with_warning(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, existing_id="E-001")

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
    )

    sheet = _output_sheet(result)
    assert result["overall_status"] == "warning"
    assert sheet.cell(row=5, column=1).value is None


def test_numeric_existing_ids_write_planned_numeric_ids(tmp_path: Path) -> None:
    result = _run_export(tmp_path, existing_id="41")

    sheet = _output_sheet(result)
    assert sheet.cell(row=5, column=1).value == "42"


def test_export_writes_markdown_and_json_reports(tmp_path: Path) -> None:
    result = _run_export(tmp_path)

    markdown_path, json_path = write_workbook_export_report(
        result,
        tmp_path / "reports",
        written_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    assert markdown_path.name == "workbook_export_2026-05-25_153000.md"
    assert json.loads(json_path.read_text(encoding="utf-8"))["rows_exported"] == 1


def test_export_writes_to_change_log(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path)

    apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
        applied_at=datetime(2026, 5, 25, 15, 30, 0),
    )

    rows = _read_rows(queue_dir / "change_log.csv")
    assert rows[0]["operation"] == "apply_approved_to_workbook"
    assert rows[0]["status"] == "success"


def test_proposal_id_filters_export_to_one_proposal(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, extra_approved=True)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
        proposal_id="PROP0002",
    )

    sheet = _output_sheet(result)
    assert result["rows_exported"] == 1
    assert sheet.cell(row=5, column=3).value == "Second approved evidence."


def test_source_id_filters_export_to_one_source(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, extra_approved=True)

    result = apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
        source_id="SRC0002",
    )

    sheet = _output_sheet(result)
    assert result["rows_exported"] == 1
    assert sheet.cell(row=5, column=5).value == "Second Source"


def _run_export(tmp_path: Path, *, existing_id: str = "1") -> dict:
    config, queue_dir, workbook_path = _setup_export_fixture(tmp_path, existing_id=existing_id)
    return apply_approved_to_workbook(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        backups_dir=tmp_path / "backups",
        outputs_dir=tmp_path / "outputs",
        applied_at=datetime(2026, 5, 25, 15, 30, 0),
    )


def _setup_export_fixture(
    tmp_path: Path,
    *,
    approved_overrides: dict[str, str] | None = None,
    existing_id: str = "1",
    omit_notes: bool = False,
    extra_approved: bool = False,
    stale_blank_row: int | None = None,
    formula_row: int | None = None,
) -> tuple[dict, Path, Path]:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    init_queues(queue_dir, config)
    workbook_path = tmp_path / "workbook.xlsx"
    _create_workbook(workbook_path, existing_id=existing_id, omit_notes=omit_notes, stale_blank_row=stale_blank_row, formula_row=formula_row)
    _append_queue_row(queue_dir / "source_dossiers.csv", "source_dossiers", {"source_id": "SRC0001"})
    _append_queue_row(queue_dir / "extracted_claims.csv", "extracted_claims", {"claim_id": "C001", "source_id": "SRC0001", "claim_text": "Claim."})
    _append_queue_row(queue_dir / "proposed_updates.csv", "proposed_updates", {"proposal_id": "PROP0001", "claim_id": "C001", "source_id": "SRC0001"})
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row(approved_overrides))
    if extra_approved:
        _append_queue_row(queue_dir / "source_dossiers.csv", "source_dossiers", {"source_id": "SRC0002"})
        _append_queue_row(queue_dir / "extracted_claims.csv", "extracted_claims", {"claim_id": "C002", "source_id": "SRC0002", "claim_text": "Claim two."})
        _append_queue_row(queue_dir / "proposed_updates.csv", "proposed_updates", {"proposal_id": "PROP0002", "claim_id": "C002", "source_id": "SRC0002"})
        _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row({"proposal_id": "PROP0002", "claim_id": "C002", "source_id": "SRC0002", "evidence_argument": "Second approved evidence.", "source_book": "Second Source"}))
    return config, queue_dir, workbook_path


def _create_workbook(
    path: Path,
    *,
    existing_id: str,
    omit_notes: bool,
    stale_blank_row: int | None = None,
    formula_row: int | None = None,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence Log"
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "EC Numeric",
        "EC Factor",
        "EC Log Factor",
    ]
    if not omit_notes:
        headers.append("Notes")
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index, value=header)
    sheet.cell(row=4, column=1, value=existing_id)
    sheet.cell(row=4, column=3, value="Existing evidence.")
    sheet.cell(row=4, column=16, value="=A4")
    if formula_row:
        sheet.cell(row=formula_row, column=1, value=str(int(existing_id) + 1) if str(existing_id).isdigit() else existing_id)
        sheet.cell(row=formula_row, column=3, value="Last meaningful evidence.")
        sheet.cell(row=formula_row, column=16, value=f"=A{formula_row}")
    if stale_blank_row:
        sheet.row_dimensions[stale_blank_row].height = 15
    workbook.save(path)


def _approved_row(overrides: dict[str, str] | None = None) -> dict[str, str]:
    row = {
        "proposal_id": "PROP0001",
        "claim_id": "C001",
        "source_id": "SRC0001",
        "evidence_argument": "Approved evidence.",
        "category": "Example",
        "source_book": "Example Source",
        "approved_weight_0_5": "3",
        "EC_MI5": "Likely / probable",
        "PC_MI5": "Roughly even chance",
        "notes": "Approved note.",
        "approved_by": "Eric",
        "approved_date": "2026-05-25",
    }
    if overrides:
        row.update(overrides)
    return row


def _append_queue_row(path: Path, queue_name: str, row: dict[str, str]) -> None:
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUEUE_SCHEMAS[queue_name])
        writer.writerow({header: row.get(header, "") for header in QUEUE_SCHEMAS[queue_name]})


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _output_sheet(result: dict):
    workbook = load_workbook(result["output_workbook_path"], data_only=False)
    return workbook["Evidence Log"]
