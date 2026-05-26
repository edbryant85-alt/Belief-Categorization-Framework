from __future__ import annotations

import csv
import json
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from belief_dashboard.config import load_config
from belief_dashboard.export_verification import (
    latest_output_workbook,
    verify_workbook_export,
    write_export_verification_reports,
)
from belief_dashboard.manual_imports import queue_summary
from belief_dashboard.queues import init_queues, migrate_approved_updates_schema, validate_queues
from belief_dashboard.schemas import QUEUE_SCHEMAS


def test_latest_output_workbook_returns_most_recent_xlsx(tmp_path: Path) -> None:
    older = tmp_path / "older.xlsx"
    newer = tmp_path / "newer.xlsx"
    older.write_text("old", encoding="utf-8")
    time.sleep(0.01)
    newer.write_text("new", encoding="utf-8")

    assert latest_output_workbook(tmp_path) == newer


def test_verification_passes_when_expected_approved_rows_are_present(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path)

    result = verify_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "pass"
    assert result["matching_exported_rows_found"] == 1


def test_verification_fails_when_approved_row_is_missing(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path, omit_exported_row=True)

    result = verify_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert result["missing_exported_rows"] == 1


def test_verification_fails_when_mapped_value_mismatches(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path, workbook_overrides={"Category": "Wrong"})

    result = verify_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert result["value_mismatches"] == 1


def test_verification_detects_missing_trace_metadata(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path, notes="Trace: claim_id=C001; source_id=SRC0001")

    result = verify_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert any("exported row not found by trace metadata" in error for error in result["errors"])


def test_verification_reports_formula_concerns_when_expected_formula_is_blank(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path, blank_formula=True)

    result = verify_workbook_export(workbook_path, queue_dir / "approved_updates.csv", queue_dir, config)

    assert result["overall_status"] == "fail"
    assert result["formula_concerns"] == 1


def test_verification_writes_markdown_and_json_reports(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path)
    result = verify_workbook_export(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        verified_at=datetime(2026, 5, 26, 12, 0, 0),
    )

    markdown_path, json_path = write_export_verification_reports(
        result,
        tmp_path / "reports",
        written_at=datetime(2026, 5, 26, 12, 0, 0),
    )

    assert markdown_path.name == "export_verification_2026-05-26_120000.md"
    assert json.loads(json_path.read_text(encoding="utf-8"))["overall_status"] == "pass"


def test_mark_exported_updates_approved_queue_rows_only_when_verification_succeeds(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path)
    result = verify_workbook_export(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        mark_exported=True,
        verified_at=datetime(2026, 5, 26, 12, 0, 0),
    )
    _markdown, json_path = write_export_verification_reports(result, tmp_path / "reports")

    rows = _read_rows(queue_dir / "approved_updates.csv")
    assert result["approved_rows_marked_exported"] is True
    assert rows[0]["export_status"] == "exported"
    assert rows[0]["exported_at"] == "2026-05-26T12:00:00"
    assert rows[0]["exported_workbook"] == str(workbook_path)
    assert rows[0]["export_verification_report"] == str(json_path)


def test_mark_exported_does_not_update_when_verification_fails(tmp_path: Path) -> None:
    config, queue_dir, workbook_path = _setup_verification_fixture(tmp_path, workbook_overrides={"Category": "Wrong"})

    result = verify_workbook_export(
        workbook_path,
        queue_dir / "approved_updates.csv",
        queue_dir,
        config,
        mark_exported=True,
    )

    rows = _read_rows(queue_dir / "approved_updates.csv")
    assert result["approved_rows_marked_exported"] is False
    assert rows[0]["export_status"] == ""


def test_expanded_approved_queue_schema_is_initialized_correctly(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"

    init_queues(queue_dir, config)

    assert _header(queue_dir / "approved_updates.csv") == QUEUE_SCHEMAS["approved_updates"]


def test_existing_approved_queue_rows_can_be_migrated_to_export_tracking_fields(tmp_path: Path) -> None:
    path = tmp_path / "approved_updates.csv"
    legacy_headers = QUEUE_SCHEMAS["approved_updates"][:-4]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=legacy_headers)
        writer.writeheader()
        writer.writerow({"proposal_id": "PROP0001", "claim_id": "C001", "source_id": "SRC0001"})

    migrated = migrate_approved_updates_schema(path)

    rows = _read_rows(path)
    assert migrated is True
    assert _header(path) == QUEUE_SCHEMAS["approved_updates"]
    assert rows[0]["proposal_id"] == "PROP0001"
    assert rows[0]["export_status"] == ""


def test_validate_queues_passes_with_expanded_approved_schema(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    init_queues(queue_dir, config)
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row({"export_status": "exported"}))

    result = validate_queues(queue_dir, config)

    assert result["overall_status"] == "pass"


def test_queue_summary_includes_exported_and_not_exported_approved_counts(tmp_path: Path) -> None:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    init_queues(queue_dir, config)
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row())
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row({"proposal_id": "PROP0002", "export_status": "exported"}))

    summary = queue_summary(queue_dir, config)

    assert summary["approved_updates_export_tracking"] == {
        "total": 2,
        "exported": 1,
        "not_exported": 1,
    }


def _setup_verification_fixture(
    tmp_path: Path,
    *,
    omit_exported_row: bool = False,
    workbook_overrides: dict[str, str] | None = None,
    notes: str | None = None,
    blank_formula: bool = False,
) -> tuple[dict, Path, Path]:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    workbook_path = tmp_path / "output.xlsx"
    init_queues(queue_dir, config)
    _append_queue_row(queue_dir / "approved_updates.csv", "approved_updates", _approved_row())
    _create_output_workbook(workbook_path, omit_exported_row=omit_exported_row, workbook_overrides=workbook_overrides or {}, notes=notes, blank_formula=blank_formula)
    return config, queue_dir, workbook_path


def _create_output_workbook(
    path: Path,
    *,
    omit_exported_row: bool,
    workbook_overrides: dict[str, str],
    notes: str | None,
    blank_formula: bool,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence Log"
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "EC Numeric",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=3, column=column_index, value=header)
    sheet.cell(row=4, column=16, value="=A4")
    if not omit_exported_row:
        values = {
            "Date": "2026-05-26",
            "Evidence / Argument": "Approved evidence.",
            "Category": "Example",
            "Source / Book": "Example Source",
            "Weight 0-5": "3",
            "EC MI5": "Likely / probable",
            "PC MI5": "Roughly even chance",
            "Notes": notes or "Original note: Approved note.\nTrace: proposal_id=PROP0001; claim_id=C001; source_id=SRC0001; approved_by=Eric; approved_date=2026-05-26",
            "EC Numeric": "" if blank_formula else "=A5",
        }
        values.update(workbook_overrides)
        for column_index, header in enumerate(headers, start=1):
            if header in values:
                sheet.cell(row=5, column=column_index, value=values[header])
    workbook.save(path)


def _approved_row(overrides: dict[str, str] | None = None) -> dict[str, str]:
    row = {
        "proposal_id": "PROP0001",
        "claim_id": "C001",
        "source_id": "SRC0001",
        "evidence_argument": "Approved evidence.",
        "category": "Example",
        "source_book": "Example Source",
        "approved_weight_0_5": "3",
        "EC_MI5": "Likely / probable",
        "PC_MI5": "Roughly even chance",
        "notes": "Approved note.",
        "approved_by": "Eric",
        "approved_date": "2026-05-26",
    }
    if overrides:
        row.update(overrides)
    return row


def _append_queue_row(path: Path, queue_name: str, row: dict[str, str]) -> None:
    with path.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUEUE_SCHEMAS[queue_name])
        writer.writerow({header: row.get(header, "") for header in QUEUE_SCHEMAS[queue_name]})


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _header(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return next(csv.reader(handle))
