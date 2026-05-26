from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook

from belief_dashboard.config import load_config
from belief_dashboard.history import (
    current_workbook_status,
    export_history,
    list_promoted_archives,
    promotion_history,
    verification_history,
)
from belief_dashboard.queues import init_queues
from belief_dashboard.workbook_recovery import rollback_workbook, write_workbook_rollback_reports


ROLLED_BACK_AT = datetime.now().replace(microsecond=0) + timedelta(days=1)


def test_current_workbook_status_reports_main_workbook_status(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    _write_json(fixture["export_reports"] / "workbook_export_2026-05-26_120000.json", {"export_timestamp": "2026-05-26T12:00:00"})

    status = current_workbook_status(
        main_workbook=fixture["main"],
        outputs_dir=fixture["outputs"],
        export_reports_dir=fixture["export_reports"],
        verification_reports_dir=fixture["verification_reports"],
        promotion_reports_dir=fixture["promotion_reports"],
        recovery_reports_dir=fixture["recovery_reports"],
        promoted_archive_dir=fixture["promoted_archives"],
        queue_dir=fixture["queue_dir"],
        config=fixture["config"],
    )

    assert status["main_workbook_exists"] is True
    assert status["main_workbook_modified_timestamp"]
    assert status["latest_export_report"].endswith("workbook_export_2026-05-26_120000.json")
    assert status["queue_summary"]["counts"]["approved_updates"] == 0


def test_promotion_history_reads_promotion_json_reports(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    report = fixture["promotion_reports"] / "workbook_promotion_2026-05-26_120000.json"
    _write_json(
        report,
        {
            "promotion_timestamp": "2026-05-26T12:00:00",
            "candidate_output_workbook_path": "data/outputs/output.xlsx",
            "main_workbook_path": "data/workbooks/main.xlsx",
            "archive_path": "data/backups/promoted_archives/main_pre_promotion.xlsx",
            "overall_status": "pass",
        },
    )

    history = promotion_history(fixture["promotion_reports"], fixture["queue_dir"] / "change_log.csv")

    assert history["count"] == 1
    assert history["rows"][0]["promoted_workbook"] == "data/outputs/output.xlsx"
    assert history["rows"][0]["report_path"] == str(report)


def test_export_history_reads_workbook_export_json_reports(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    _write_json(
        fixture["export_reports"] / "workbook_export_2026-05-26_120000.json",
        {
            "export_timestamp": "2026-05-26T12:00:00",
            "workbook_path": "main.xlsx",
            "backup_workbook_path": "backup.xlsx",
            "output_workbook_path": "output.xlsx",
            "rows_exported": 3,
            "overall_status": "pass",
        },
    )

    history = export_history(fixture["export_reports"], fixture["queue_dir"] / "change_log.csv")

    assert history["count"] == 1
    assert history["rows"][0]["output_workbook"] == "output.xlsx"
    assert history["rows"][0]["rows_exported"] == 3


def test_verification_history_reads_export_verification_json_reports(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    _write_json(
        fixture["verification_reports"] / "export_verification_2026-05-26_120000.json",
        {
            "verification_timestamp": "2026-05-26T12:00:00",
            "output_workbook_path": "output.xlsx",
            "overall_status": "pass",
            "approved_rows_considered": 2,
            "matching_exported_rows_found": 2,
            "missing_exported_rows": 0,
            "value_mismatches": 0,
            "formula_concerns": 0,
            "mark_exported_requested": True,
        },
    )

    history = verification_history(fixture["verification_reports"])

    assert history["count"] == 1
    assert history["rows"][0]["verified_workbook"] == "output.xlsx"
    assert history["rows"][0]["mark_exported_used"] is True


def test_list_promoted_archives_lists_promoted_archive_workbooks(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)

    result = list_promoted_archives(fixture["promoted_archives"])

    assert result["count"] == 1
    assert result["rows"][0]["archive_filename"] == fixture["archive"].name
    assert result["rows"][0]["size"] > 0


def test_rollback_dry_run_writes_no_files_and_does_not_replace_main_workbook(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    main_hash = _sha256(fixture["main"])

    result = _rollback(fixture, dry_run=True)

    assert result["overall_status"] == "pass"
    assert result["main_workbook_replaced"] is False
    assert _sha256(fixture["main"]) == main_hash
    assert not list(fixture["rollback_archives"].glob("*.xlsx"))
    assert not list(fixture["recovery_reports"].glob("*"))
    assert _read_rows(fixture["queue_dir"] / "change_log.csv") == []


def test_rollback_fails_when_archive_is_missing(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    fixture["archive"].unlink()

    result = _rollback(fixture)

    assert result["overall_status"] == "fail"
    assert any("Selected archive not found" in error for error in result["errors"])


def test_rollback_fails_when_main_workbook_is_missing(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    fixture["main"].unlink()

    result = _rollback(fixture)

    assert result["overall_status"] == "fail"
    assert any("Main workbook not found" in error for error in result["errors"])


def test_rollback_fails_when_archive_is_not_valid_workbook(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path, invalid_archive=True)

    result = _rollback(fixture)

    assert result["overall_status"] == "fail"
    assert result["archive_inspection_passed"] is False


def test_rollback_creates_archive_replaces_main_and_preserves_selected_archive(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    main_hash = _sha256(fixture["main"])
    archive_hash = _sha256(fixture["archive"])

    result = _rollback(fixture)

    rollback_archive = Path(result["rollback_archive_path"])
    assert result["overall_status"] == "pass"
    assert rollback_archive.exists()
    assert _sha256(rollback_archive) == main_hash
    assert _sha256(fixture["main"]) == archive_hash
    assert fixture["archive"].exists()
    assert _sha256(fixture["archive"]) == archive_hash


def test_rollback_writes_markdown_and_json_reports(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)
    result = _rollback(fixture)

    markdown_path, json_path = write_workbook_rollback_reports(
        result,
        fixture["recovery_reports"],
        written_at=ROLLED_BACK_AT,
    )

    assert markdown_path.exists()
    assert json_path.exists()
    assert "Workbook Rollback Report" in markdown_path.read_text(encoding="utf-8")
    assert json.loads(json_path.read_text(encoding="utf-8"))["overall_status"] == "pass"


def test_rollback_writes_to_change_log(tmp_path: Path) -> None:
    fixture = _setup_recovery_fixture(tmp_path)

    result = _rollback(fixture)

    rows = _read_rows(fixture["queue_dir"] / "change_log.csv")
    assert result["change_log_updated"] is True
    assert rows[-1]["operation"] == "rollback_workbook"
    assert rows[-1]["input_file"] == str(fixture["archive"])
    assert rows[-1]["output_file"] == str(fixture["main"])
    assert rows[-1]["status"] == "pass"


def _setup_recovery_fixture(tmp_path: Path, *, invalid_archive: bool = False) -> dict:
    config = load_config("config.yaml")
    queue_dir = tmp_path / "queues"
    main = tmp_path / "workbooks" / "main.xlsx"
    archive = tmp_path / "promoted_archives" / "main_pre_promotion.xlsx"
    fixture = {
        "config": config,
        "queue_dir": queue_dir,
        "main": main,
        "archive": archive,
        "outputs": tmp_path / "outputs",
        "export_reports": tmp_path / "reports" / "workbook_exports",
        "verification_reports": tmp_path / "reports" / "export_verification",
        "promotion_reports": tmp_path / "reports" / "workbook_promotion",
        "recovery_reports": tmp_path / "reports" / "workbook_recovery",
        "promoted_archives": tmp_path / "promoted_archives",
        "rollback_archives": tmp_path / "rollback_archives",
    }
    init_queues(queue_dir, config)
    _create_sample_workbook(main, config, marker="current-main")
    if invalid_archive:
        _create_invalid_workbook(archive)
    else:
        _create_sample_workbook(archive, config, marker="archived-main")
    for path in [
        fixture["outputs"],
        fixture["export_reports"],
        fixture["verification_reports"],
        fixture["promotion_reports"],
        fixture["recovery_reports"],
        fixture["rollback_archives"],
    ]:
        path.mkdir(parents=True, exist_ok=True)
    return fixture


def _rollback(fixture: dict, *, dry_run: bool = False) -> dict:
    return rollback_workbook(
        fixture["archive"],
        fixture["main"],
        fixture["config"],
        rollback_archive_dir=fixture["rollback_archives"],
        reports_dir=fixture["recovery_reports"],
        queue_dir=fixture["queue_dir"],
        dry_run=dry_run,
        rolled_back_at=ROLLED_BACK_AT,
    )


def _create_sample_workbook(path: Path, config: dict, *, marker: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]
    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)
    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)
    evidence.cell(row=header_row + 1, column=1, value=marker)
    evidence.cell(row=header_row + 1, column=3, value=f"{marker} evidence")
    workbook.save(path)


def _create_invalid_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.active.title = "Unexpected"
    workbook.save(path)


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))
