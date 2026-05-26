from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import Workbook

from belief_dashboard.cli import main
from belief_dashboard.config import load_config
from belief_dashboard.operator_preflight import build_operator_preflight
from belief_dashboard.queues import init_queues
from belief_dashboard.schemas import QUEUE_SCHEMAS


def test_operator_preflight_works_in_general_mode(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)

    result = build_operator_preflight(config, tmp_path, mode="general")

    assert result["operation"] == "operator_preflight"
    assert result["mode"] == "general"
    assert result["workbook"]["overall_status"] == "pass"
    assert result["queue_validation"]["overall_status"] == "pass"
    assert result["recommended_next_commands"]


def test_operator_preflight_before_export_reports_workbook_and_queue_status(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config, approved_rows=1)

    result = build_operator_preflight(config, tmp_path, mode="before-export")

    assert result["overall_status"] in {"pass", "warning"}
    assert result["workbook"]["overall_status"] == "pass"
    assert result["queue_validation"]["overall_status"] == "pass"
    assert result["queue_summary"]["approved_updates_export_tracking"]["not_exported"] == 1
    assert any("apply-approved-to-workbook --dry-run" in command for command in result["recommended_next_commands"])


def test_operator_preflight_before_verification_finds_latest_output_workbook(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)
    output = tmp_path / "outputs" / "output.xlsx"
    _write_file(output, b"output")

    result = build_operator_preflight(config, tmp_path, mode="before-verification")

    assert result["latest"]["output_workbook"]["path"] == str(output)
    assert any("verify-workbook-export" in command for command in result["recommended_next_commands"])


def test_operator_preflight_before_promotion_finds_passing_verified_output(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)
    output = tmp_path / "outputs" / "output.xlsx"
    report = tmp_path / "reports" / "export_verification" / "verification.json"
    _write_file(output, b"output")
    _write_json(report, {"overall_status": "pass", "output_workbook_path": str(output), "verification_timestamp": "2026-05-26T12:00:00"})

    result = build_operator_preflight(config, tmp_path, mode="before-promotion")

    assert result["overall_status"] == "pass"
    assert result["verified_outputs"]["count"] == 1
    assert result["command_guides"]["promote"]["dry_run_command"]
    assert any("promote-output-workbook" in command for command in result["recommended_next_commands"])


def test_operator_preflight_before_promotion_fails_when_no_passing_verification_exists(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)

    result = build_operator_preflight(config, tmp_path, mode="before-promotion")

    assert result["overall_status"] == "fail"
    assert "No passing verification report is available for promotion." in result["errors"]


def test_operator_preflight_before_rollback_finds_latest_promoted_archive(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)
    older = tmp_path / "backups" / "promoted_archives" / "older.xlsx"
    newer = tmp_path / "backups" / "promoted_archives" / "newer.xlsx"
    _write_file(older, b"older")
    _write_file(newer, b"newer")
    _touch(older, datetime(2026, 5, 26, 10, 0, 0))
    _touch(newer, datetime(2026, 5, 26, 11, 0, 0))

    result = build_operator_preflight(config, tmp_path, mode="before-rollback")

    assert result["overall_status"] == "pass"
    assert result["latest"]["promoted_archive"]["path"] == str(newer)
    assert result["command_guides"]["rollback"]["dry_run_command"]


def test_operator_preflight_before_rollback_fails_when_no_archive_exists(tmp_path: Path) -> None:
    config = _preflight_config(tmp_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)

    result = build_operator_preflight(config, tmp_path, mode="before-rollback")

    assert result["overall_status"] == "fail"
    assert "No promoted archive is available for rollback." in result["errors"]


def test_operator_preflight_format_json_returns_structured_output(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)

    exit_code = main(["operator-preflight", "--config", str(config_path), "--format", "json"])

    output = json.loads(capsys.readouterr().out)
    assert exit_code == 0
    assert output["operation"] == "operator_preflight"
    assert output["mode"] == "general"


def test_operator_preflight_save_writes_markdown_and_json_reports(tmp_path: Path, capsys) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config)

    exit_code = main(["operator-preflight", "--config", str(config_path), "--save"])

    output = capsys.readouterr().out
    reports_dir = tmp_path / "reports" / "operator_preflight"
    assert exit_code == 0
    assert "Markdown report:" in output
    assert "JSON report:" in output
    assert list(reports_dir.glob("operator_preflight_GENERAL_*.md"))
    assert list(reports_dir.glob("operator_preflight_GENERAL_*.json"))


def test_operator_preflight_does_not_modify_workbooks(tmp_path: Path) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    workbook = Path(config["workbook"]["default_path"])
    _create_sample_workbook(workbook, config)
    _setup_queues(tmp_path, config)
    before = _sha256(workbook)

    assert main(["operator-preflight", "--config", str(config_path), "--mode", "before-export"]) in {0, 1}

    assert _sha256(workbook) == before


def test_operator_preflight_does_not_modify_queue_csv_files(tmp_path: Path) -> None:
    config_path = _write_config(tmp_path)
    config = load_config(config_path)
    _create_sample_workbook(Path(config["workbook"]["default_path"]), config)
    _setup_queues(tmp_path, config, approved_rows=1)
    queue_file = tmp_path / "queues" / "approved_updates.csv"
    before = _sha256(queue_file)

    assert main(["operator-preflight", "--config", str(config_path), "--mode", "before-export"]) in {0, 1}

    assert _sha256(queue_file) == before


def _preflight_config(tmp_path: Path) -> dict:
    config = load_config("config.yaml")
    config["workbook"]["default_path"] = str(tmp_path / "workbooks" / "main.xlsx")
    config["queues"]["base_dir"] = str(tmp_path / "queues")
    config["artifact_navigation"] = {
        "reports_dir": str(tmp_path / "reports" / "artifact_navigation"),
        "reports": {
            "workbook_inspection": str(tmp_path / "reports" / "workbook_inspection"),
            "queue_validation": str(tmp_path / "reports" / "queue_validation"),
            "prompt_packets": str(tmp_path / "reports" / "prompt_packets"),
            "manual_imports": str(tmp_path / "reports" / "manual_imports"),
            "reviews": str(tmp_path / "reports" / "reviews"),
            "workbook_export_preview": str(tmp_path / "reports" / "workbook_export_preview"),
            "workbook_exports": str(tmp_path / "reports" / "workbook_exports"),
            "export_verification": str(tmp_path / "reports" / "export_verification"),
            "workbook_promotion": str(tmp_path / "reports" / "workbook_promotion"),
            "workbook_recovery": str(tmp_path / "reports" / "workbook_recovery"),
        },
        "workbooks": {
            "main": str(tmp_path / "workbooks" / "main.xlsx"),
            "outputs": str(tmp_path / "outputs"),
            "backups": str(tmp_path / "backups"),
            "promoted_archives": str(tmp_path / "backups" / "promoted_archives"),
            "rollback_archives": str(tmp_path / "backups" / "rollback_archives"),
        },
        "default_limit": 10,
    }
    config["command_composition"] = {
        "reports_dir": str(tmp_path / "reports" / "command_guides"),
        "default_include_dry_run_first": True,
        "quote_paths": True,
    }
    config["operator_preflight"] = {
        "reports_dir": str(tmp_path / "reports" / "operator_preflight"),
        "default_mode": "general",
    }
    for directory in config["artifact_navigation"]["reports"].values():
        Path(directory).mkdir(parents=True, exist_ok=True)
    for directory in [
        config["artifact_navigation"]["reports_dir"],
        config["artifact_navigation"]["workbooks"]["outputs"],
        config["artifact_navigation"]["workbooks"]["promoted_archives"],
        config["artifact_navigation"]["workbooks"]["rollback_archives"],
        config["command_composition"]["reports_dir"],
        config["operator_preflight"]["reports_dir"],
    ]:
        Path(directory).mkdir(parents=True, exist_ok=True)
    Path(config["workbook"]["default_path"]).parent.mkdir(parents=True, exist_ok=True)
    return config


def _write_config(tmp_path: Path) -> Path:
    config = _preflight_config(tmp_path)
    config_path = tmp_path / "config.yaml"
    config_path.write_text(yaml.safe_dump(config), encoding="utf-8")
    return config_path


def _setup_queues(tmp_path: Path, config: dict, *, approved_rows: int = 0) -> None:
    queue_dir = tmp_path / "queues"
    init_queues(queue_dir, config)
    if approved_rows:
        approved_path = queue_dir / config["queues"]["files"]["approved_updates"]
        with approved_path.open("a", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=QUEUE_SCHEMAS["approved_updates"])
            for index in range(approved_rows):
                writer.writerow(
                    {
                        "proposal_id": f"PROP{index + 1:04d}",
                        "claim_id": f"CLM{index + 1:04d}",
                        "source_id": "SRC0001",
                        "evidence_argument": "Evidence",
                        "category": "Test",
                        "source_book": "Source",
                    }
                )


def _create_sample_workbook(path: Path, config: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = config["workbook"]["expected_sheets"][0]
    for sheet_name in config["workbook"]["expected_sheets"][1:]:
        workbook.create_sheet(sheet_name)
    evidence = workbook[config["workbook"]["evidence_log"]["sheet_name"]]
    header_row = config["workbook"]["evidence_log"]["header_row"]
    headers = [
        "ID",
        "Date",
        "Evidence / Argument",
        "Category",
        "Source / Book",
        "Weight 0-5",
        "EC MI5",
        "PC MI5",
        "PT MI5",
        "CT MI5",
        "MT MI5",
        "IS MI5",
        "MS MI5",
        "HC MI5",
        "N MI5",
        "Notes",
    ]
    for column_index, header in enumerate(headers, start=1):
        evidence.cell(row=header_row, column=column_index, value=header)
    workbook.save(path)


def _write_file(path: Path, content: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)


def _write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")


def _touch(path: Path, value: datetime) -> None:
    timestamp = value.timestamp()
    path.touch()
    import os

    os.utime(path, (timestamp, timestamp))


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
